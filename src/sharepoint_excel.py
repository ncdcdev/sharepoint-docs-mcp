"""
SharePoint Excel解析モジュール（ダウンロード+openpyxl方式）
"""

import difflib
import json
import logging
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell import Cell

from src.config import config
from src.excel import (
    ExcelMergedCellHandler,
    ExcelPaneManager,
    ExcelRangeCalculator,
    ExcelStyleExtractor,
)

logger = logging.getLogger(__name__)


class SharePointExcelParser:
    """SharePoint Excelファイル解析クライアント"""

    def __init__(self, download_client):
        """
        Args:
            download_client: download_file(file_path) -> bytes メソッドを持つクライアント
        """
        self.download_client = download_client

    def search_cells(
        self,
        file_path: str,
        query: str,
        sheet_name: str | None = None,
        include_row_data: bool = False,
    ) -> str:
        """
        セル内容を検索して該当位置を返す

        Args:
            file_path: Excelファイルのパス
            query: 検索キーワード
            sheet_name: 検索対象シート名（指定時はまずそのシートを検索し、マッチ0件なら全シート検索にフォールバック）

        Returns:
            JSON文字列（マッチしたセルの位置情報）
        """
        logger.info(
            f"Searching cells in Excel file: {file_path} (query={query}, sheet={sheet_name})"
        )

        try:
            # ファイルをダウンロード
            file_bytes = self.download_client.download_file(file_path)
            logger.info(f"Downloaded {len(file_bytes)} bytes")

            # BytesIOでメモリ上に展開
            file_stream = BytesIO(file_bytes)

            # openpyxlで読み込み
            workbook = load_workbook(file_stream, data_only=False, rich_text=True)

            matches = []
            warnings = []

            # sheet_name 指定がある場合はそのシートを優先して検索
            if sheet_name:
                if sheet_name in workbook.sheetnames:
                    self._scan_sheet(
                        workbook[sheet_name],
                        sheet_name,
                        query,
                        matches,
                        include_row_data,
                    )

                    # マッチが無ければ全シート走査にフォールバック
                    if len(matches) == 0:
                        for sn in workbook.sheetnames:
                            if sn == sheet_name:
                                continue
                            self._scan_sheet(
                                workbook[sn], sn, query, matches, include_row_data
                            )
                else:
                    # sheet_name が存在しない場合は「指定なし」と同じ扱いで全シート検索
                    warnings.append(
                        f"Sheet '{sheet_name}' not found. Searching all sheets instead."
                    )
                    for sn in workbook.sheetnames:
                        self._scan_sheet(
                            workbook[sn], sn, query, matches, include_row_data
                        )
            else:
                # 全シート検索
                for sn in workbook.sheetnames:
                    self._scan_sheet(workbook[sn], sn, query, matches, include_row_data)

            logger.info(f"Found {len(matches)} matches for query '{query}'")

            result = {
                "file_path": file_path,
                "mode": "search",
                "query": query,
                "match_count": len(matches),
                "matches": matches,
            }
            if warnings:
                result["warnings"] = warnings

            return json.dumps(result, ensure_ascii=False, indent=2)

        except Exception as e:
            logger.error(f"Failed to search cells in Excel file: {str(e)}")
            raise

    def parse_to_json(
        self,
        file_path: str,
        sheet_name: str | None = None,
        cell_range: str | None = None,
        include_frozen_rows: bool = True,
        include_cell_styles: bool = False,
        expand_axis_range: bool = False,
    ) -> str:
        """
        Excelファイルを解析してJSON形式で返す

        Args:
            file_path: Excelファイルのパス
            sheet_name: 特定シートのみ取得（Noneで全シート）
            cell_range: セル範囲指定（例: "A1:D10"）
            include_frozen_rows: cell_range指定時に固定行（ヘッダー）を自動追加
                True（デフォルト）: frozen_rowsで指定された行を自動的に取得
                False: 指定されたcell_rangeのみを取得
            include_cell_styles: セルの色・サイズ情報（default: false）
                色分けデータ抽出時のみ使用。トークン消費+約20%
            expand_axis_range: 単一列/行指定時に開始側を自動拡張（default: false）
                True: 例 "J50:J100" → "J1:J100"（行1に拡張）
                False: 指定範囲をそのまま使用

        Returns:
            JSON文字列
            - 各セルのデータ: value（値）、coordinate（座標）
            - 構造情報: シート名、dimensions（シート全体のセル範囲、例: "A1:D10"）
            - 構造情報: frozen_rows（固定行数）、frozen_cols（固定列数）
            - 条件付き構造情報: freeze_panes（存在する場合）、merged_ranges（結合セルが存在する場合）
            - スタイル情報（include_cell_styles=Trueの場合）: fill（背景色）、width（列幅）、height（行高さ）
        """
        logger.info(
            f"Parsing Excel file: {file_path} (sheet={sheet_name}, range={cell_range})"
        )

        try:
            # ファイルをダウンロード
            file_bytes = self.download_client.download_file(file_path)
            logger.info(f"Downloaded {len(file_bytes)} bytes")

            # BytesIOでメモリ上に展開
            file_stream = BytesIO(file_bytes)

            # openpyxlで読み込み（data_only=Falseで数式も取得）
            workbook = load_workbook(file_stream, data_only=False, rich_text=True)

            # シートリストを取得
            sheet_resolution: dict[str, Any] | None = None

            if sheet_name:
                resolved, candidates = self._resolve_sheet_name(
                    workbook.sheetnames, sheet_name
                )

                if resolved:
                    sheets_to_parse = [resolved]
                    if resolved != sheet_name:
                        sheet_resolution = {
                            "status": "resolved",
                            "requested": sheet_name,
                            "resolved": resolved,
                        }
                else:
                    # sheet_name が解決できない場合
                    # cell_range が指定されていれば、範囲が限定されるので全シートにフォールバック（やりすぎを避けつつ取りこぼし防止）
                    if cell_range:
                        sheets_to_parse = workbook.sheetnames
                        sheet_resolution = {
                            "status": "fallback_all_sheets",
                            "requested": sheet_name,
                            "resolved": None,
                            "candidates": candidates,
                            "reason": "sheet not found; fallback to all sheets because cell_range is specified",
                        }
                    else:
                        # ここで例外にせず、空の sheets を返して候補を出す（LLMが次の手を打てる）
                        sheets_to_parse = []
                        sheet_resolution = {
                            "status": "not_found",
                            "requested": sheet_name,
                            "resolved": None,
                            "candidates": candidates,
                        }
            else:
                sheets_to_parse = workbook.sheetnames

            # シートを解析
            result = {
                "file_path": file_path,
                "sheets": [],
            }

            # nullでない場合のみ追加
            if sheet_name is not None:
                result["requested_sheet"] = sheet_name
            if cell_range is not None:
                result["requested_range"] = cell_range

            if sheet_resolution:
                result["sheet_resolution"] = sheet_resolution
                result["available_sheets"] = workbook.sheetnames
                if sheet_resolution.get("status") != "resolved":
                    result["warning"] = (
                        "requested sheet_name was not found or ambiguous"
                    )

            for name in sheets_to_parse:
                sheet = workbook[name]
                sheet_data = self._parse_sheet(
                    sheet,
                    cell_range,
                    include_frozen_rows,
                    include_cell_styles,
                    expand_axis_range,
                )
                result["sheets"].append(sheet_data)

            logger.info(f"Parsed {len(result['sheets'])} sheets")
            return json.dumps(result, ensure_ascii=False, indent=2)

        except Exception as e:
            logger.error(f"Failed to parse Excel file: {str(e)}")
            raise

    def _resolve_sheet_name(
        self,
        sheetnames: list[str],
        requested: str,
    ) -> tuple[str | None, list[str]]:
        """
        sheet_name を解決する
        - 完全一致 → そのまま
        - trim + casefold 一致が 1件 → そのシート名に解決
        - それ以外 → None と候補（曖昧一致 or 類似名）を返す
        """
        if requested in sheetnames:
            return (requested, [])

        req_norm = requested.strip().casefold()

        norm_map: dict[str, list[str]] = {}
        for sn in sheetnames:
            key = sn.strip().casefold()
            norm_map.setdefault(key, []).append(sn)

        # 正規化一致
        if req_norm in norm_map:
            candidates = norm_map[req_norm]
            if len(candidates) == 1:
                return (candidates[0], [])
            # 同一正規化で複数（曖昧）
            return (None, candidates)

        # 類似名候補（表示用）
        suggestions = difflib.get_close_matches(requested, sheetnames, n=3, cutoff=0.6)
        return (None, suggestions)

    def _scan_sheet(
        self,
        sheet,
        sheet_name_for_result: str,
        query: str,
        matches: list[dict[str, Any]],
        include_row_data: bool = False,
    ) -> None:
        """
        シート内のセルを走査してqueryに一致するセルをmatchesに追加する
        """
        # 空シートを避ける意図
        if sheet.dimensions:
            # パフォーマンスのため_cellsを優先し、無い場合は公開APIにフォールバック
            # 注意: _cellsはopenpyxlのプライベート属性のため、将来のバージョンで変更される可能性があります。
            # その場合はiter_rows()を使用するフォールバックロジックが動作します。
            if hasattr(sheet, "_cells"):
                # 実在セルのみを走査（高速）
                # まずマッチを収集（_cellsのイテレーション中にsheetアクセスすると辞書が変わるため）
                new_matches: list[dict[str, Any]] = []
                for cell in sheet._cells.values():
                    if cell.value is not None:
                        cell_value_str = str(cell.value)
                        if query in cell_value_str:
                            new_matches.append(
                                {
                                    "sheet": sheet_name_for_result,
                                    "coordinate": cell.coordinate,
                                    "value": self._serialize_value(cell.value),
                                    "_row": cell.row,
                                }
                            )
                # イテレーション完了後に行データを取得
                for match in new_matches:
                    row_num = match.pop("_row")
                    if include_row_data:
                        match["row_data"] = self._get_row_data(sheet, row_num)
                    matches.append(match)
            else:
                # openpyxl公開APIを使用（互換性確保）
                for row in sheet.iter_rows(values_only=False):
                    for cell in row:
                        if cell.value is not None:
                            cell_value_str = str(cell.value)
                            if query in cell_value_str:
                                match = {
                                    "sheet": sheet_name_for_result,
                                    "coordinate": cell.coordinate,
                                    "value": self._serialize_value(cell.value),
                                }
                                if include_row_data:
                                    match["row_data"] = [
                                        {
                                            "coordinate": c.coordinate,
                                            "value": self._serialize_value(c.value),
                                        }
                                        for c in row
                                        if c.value is not None
                                    ]
                                matches.append(match)

    def _get_row_data(self, sheet, row_num: int) -> list[dict[str, Any]]:
        """
        指定行の非nullセルデータをリストとして返す

        Args:
            sheet: openpyxl Worksheet
            row_num: 行番号

        Returns:
            非nullセルの [{coordinate, value}, ...] リスト
        """
        row_cells = sheet[row_num]
        # 単一列シートではCellオブジェクト単体が返される場合がある
        if isinstance(row_cells, Cell):
            row_cells = (row_cells,)
        return [
            {
                "coordinate": c.coordinate,
                "value": self._serialize_value(c.value),
            }
            for c in row_cells
            if c.value is not None
        ]

    def _parse_sheet(
        self,
        sheet,
        cell_range: str | None = None,
        include_frozen_rows: bool = True,
        include_cell_styles: bool = False,
        expand_axis_range: bool = False,
    ) -> dict[str, Any]:
        """
        シートを解析してdict形式で返す

        Args:
            sheet: openpyxl Worksheet
            cell_range: セル範囲指定（例: "A1:D10"）
            include_frozen_rows: cell_range指定時に固定行（ヘッダー）を自動追加
            include_cell_styles: セルのスタイル情報を含めるか
            expand_axis_range: 単一列/行指定時に開始側を自動拡張

        Returns:
            シートデータのdict
        """
        sheet_data = {
            "name": sheet.title,
        }

        # dimensionsがNoneでない場合のみ追加
        if sheet.dimensions:
            sheet_data["dimensions"] = str(sheet.dimensions)

        # freeze_panes情報の取得と検証（ヘルパークラスを使用）
        frozen_rows, frozen_cols = ExcelPaneManager.get_frozen_panes(sheet)

        # frozen_rows検証（DoS対策）
        # frozen_rowsは補助的なメタ情報なので、上限超過時はリセットして処理を続行
        frozen_rows_ignored = False
        is_valid, validated_frozen_rows = ExcelPaneManager.validate_frozen_rows(
            frozen_rows, config.excel_max_frozen_rows
        )
        if not is_valid:
            logger.warning(
                "固定行数が上限(%d)を超えたため、freeze_panes情報を無視します。"
                "ファイルの解析は続行されますが、ヘッダー自動追加機能は利用できません。"
                " (frozen_rows=%d, sheet=%s)",
                config.excel_max_frozen_rows,
                frozen_rows,
                sheet.title,
            )
            frozen_rows_ignored = True
            frozen_rows = validated_frozen_rows
            frozen_cols = 0  # freeze_panes全体を無視

        if frozen_rows > 0 or frozen_cols > 0:
            sheet_data["freeze_panes"] = ExcelPaneManager.format_freeze_panes(
                frozen_rows, frozen_cols
            )
        sheet_data["frozen_rows"] = frozen_rows
        sheet_data["frozen_cols"] = frozen_cols

        # frozen_rows=0 かつ cell_range指定時、expand_axis_range=Falseの場合のみ警告
        # expand_axis_range=Trueの場合は1行目/A列が含まれるため警告不要
        if frozen_rows == 0 and cell_range and not expand_axis_range:
            status, note = (
                (
                    "ignored_due_to_limit",
                    "This sheet has frozen rows but they exceed the limit and were ignored. Headers are not automatically included.",
                )
                if frozen_rows_ignored
                else (
                    "no_frozen_rows",
                    "This sheet has no frozen rows. Headers are not automatically included.",
                )
            )
            sheet_data["header_detection"] = {
                "status": status,
                "frozen_rows": 0,
                "note": note,
                "suggestions": [
                    "If headers are needed, read 'A1:Z5' to check header structure",
                    "Or retry with expand_axis_range=True to include row 1 (for columns) or column A (for rows)",
                ],
            }

        # セル範囲の正規化・拡張（cell_rangeがある場合）
        # マージセル情報のキャッシュに使用するため、先に計算する（ヘルパークラスを使用）
        effective_range_for_merge = None
        header_range = None  # ヘッダー範囲（再利用のため事前に初期化）
        if cell_range:
            sheet_data["requested_range"] = cell_range
            max_row = sheet.max_row or 1
            effective_range = ExcelRangeCalculator.normalize_column_range(
                cell_range, max_row
            )
            if expand_axis_range:
                expanded_range = ExcelRangeCalculator.expand_axis_range(effective_range)
                if expanded_range != effective_range:
                    logger.info(
                        "Expanded axis range '%s' -> '%s' (sheet=%s)",
                        effective_range,
                        expanded_range,
                        sheet.title,
                    )
                    effective_range = expanded_range

            if effective_range != cell_range:
                logger.info(
                    "Normalized column range '%s' -> '%s' (sheet=%s)",
                    cell_range,
                    effective_range,
                    sheet.title,
                )
            sheet_data["effective_range"] = effective_range
            effective_range_for_merge = effective_range

            # ヘッダー自動追加の場合、マージセルキャッシュにもヘッダー範囲を含める
            if include_frozen_rows and frozen_rows > 0:
                header_range = ExcelRangeCalculator.calculate_header_range(
                    effective_range, frozen_rows
                )
                if header_range:
                    # ヘッダー範囲とデータ範囲を結合した範囲を計算
                    effective_range_for_merge = ExcelRangeCalculator.merge_ranges(
                        header_range, effective_range
                    )

        # データサイズ検証（DoS対策）（ヘルパークラスを使用）
        # マージセルキャッシュ構築前に検証することで、巨大な範囲によるメモリ枯渇を防ぐ
        all_rows = []

        if cell_range:
            # effective_rangeは既に計算済み
            # データサイズ検証（DoS対策）
            range_rows, range_cols = ExcelRangeCalculator.calculate_range_size(
                effective_range
            )
            if (
                range_rows > config.excel_max_data_rows
                or range_cols > config.excel_max_data_cols
            ):
                raise ValueError(
                    f"データサイズ({range_rows}行 × {range_cols}列)が上限"
                    f"({config.excel_max_data_rows}行 × {config.excel_max_data_cols}列)を超えています。"
                    f"cell_rangeパラメータで必要な範囲を指定してください。"
                    f"例: cell_range='A1:Z1000'"
                )

        elif sheet.dimensions:
            # シート全体を取得
            # データサイズ検証（DoS対策）
            sheet_rows, sheet_cols = ExcelRangeCalculator.calculate_range_size(
                sheet.dimensions
            )
            if (
                sheet_rows > config.excel_max_data_rows
                or sheet_cols > config.excel_max_data_cols
            ):
                raise ValueError(
                    f"シート全体のサイズ({sheet_rows}行 × {sheet_cols}列)が上限"
                    f"({config.excel_max_data_rows}行 × {config.excel_max_data_cols}列)を超えています。"
                    f"cell_rangeパラメータで必要な範囲を指定してください。"
                    f"例: cell_range='A1:Z1000'"
                )

        # データサイズ検証後にマージセル情報をキャッシュ（ヘルパークラスを使用）
        # 計算済みのeffective_range(effective_range_for_merge)を渡してキャッシュを構築し、
        # 戻り値としてmerged_ranges(結合セル範囲の一覧)を取得することで重複計算を回避
        merged_cell_map, merged_anchor_value_map, merged_ranges = (
            ExcelMergedCellHandler.build_merged_cell_cache(
                sheet, effective_range_for_merge, self._serialize_value
            )
        )

        # ここは「結合セルがある時だけ」返す
        if merged_ranges:
            sheet_data["merged_ranges"] = merged_ranges

        # セルサイズのキャッシュを構築（ヘルパークラスを使用）
        col_widths: dict[str, float] | None = None
        row_heights: dict[int, float] | None = None
        if include_cell_styles:
            col_widths, row_heights = ExcelStyleExtractor.build_cell_size_cache(sheet)

        # データ取得
        if cell_range:
            # ヘッダー自動追加（include_frozen_rows=Trueの場合）
            # header_rangeは既に計算済みなので再利用
            if header_range:
                # ヘッダー範囲を取得
                header_data = sheet[header_range]
                header_rows = self._normalize_range_data(header_data)
                all_rows.extend(
                    self._parse_rows(
                        header_rows,
                        include_cell_styles,
                        merged_cell_map,
                        merged_anchor_value_map,
                        col_widths,
                        row_heights,
                    )
                )

            # 通常のセル範囲取得（データ範囲）
            range_data = sheet[effective_range]
            rows_to_process = self._normalize_range_data(range_data)
            all_rows.extend(
                self._parse_rows(
                    rows_to_process,
                    include_cell_styles,
                    merged_cell_map,
                    merged_anchor_value_map,
                    col_widths,
                    row_heights,
                )
            )

        elif sheet.dimensions:
            # 全データを取得
            rows_to_process = tuple(sheet.iter_rows())

            if rows_to_process:
                all_rows.extend(
                    self._parse_rows(
                        rows_to_process,
                        include_cell_styles,
                        merged_cell_map,
                        merged_anchor_value_map,
                        col_widths,
                        row_heights,
                    )
                )

        sheet_data["rows"] = all_rows
        return sheet_data

    def _parse_cell(
        self,
        cell,
        include_cell_styles: bool = False,
        merged_cell_map: dict[str, str] | None = None,
        merged_anchor_value_map: dict[str, Any] | None = None,
        col_widths: dict[str, float] | None = None,
        row_heights: dict[int, float] | None = None,
    ) -> dict[str, Any]:
        """
        セルを解析してdict形式で返す

        Args:
            cell: openpyxl Cell
            include_cell_styles: セルのスタイル情報を含めるか（デフォルト: False）
            merged_cell_map: マージセル座標からマージ範囲へのマップ（パフォーマンス最適化用）
            merged_anchor_value_map: マージ範囲 -> アンカー値 のマップ（結合セルの値埋め用）
            col_widths: 列幅のキャッシュ（パフォーマンス最適化用）
            row_heights: 行高さのキャッシュ（パフォーマンス最適化用）

        Returns:
            セルデータのdict
        """
        # 基本情報（常に含む）
        cell_data = {
            "value": self._serialize_value(cell.value),
            "coordinate": cell.coordinate,
        }

        # セル結合情報（構造理解に必要）
        if merged_cell_map and cell.coordinate in merged_cell_map:
            merged_range_str = merged_cell_map[cell.coordinate]
            range_start = merged_range_str.split(":")[0]
            cell_data["merged"] = {
                "range": merged_range_str,
                "is_top_left": cell.coordinate == range_start,
            }

            # 結合セル内の空セルにも value を埋める（propagate）
            if cell_data["value"] is None and merged_anchor_value_map:
                anchor_value = merged_anchor_value_map.get(merged_range_str)
                if anchor_value is not None:
                    cell_data["value"] = anchor_value

        # スタイル情報（include_cell_styles=Trueの場合のみ）（ヘルパークラスを使用）
        if include_cell_styles:
            styles = ExcelStyleExtractor.extract_cell_styles(
                cell, col_widths, row_heights
            )
            # スタイル情報をcell_dataにマージ
            cell_data.update(styles)

        return cell_data

    def _parse_rows(
        self,
        rows: tuple[tuple[Cell, ...], ...],
        include_cell_styles: bool = False,
        merged_cell_map: dict[str, str] | None = None,
        merged_anchor_value_map: dict[str, Any] | None = None,
        col_widths: dict[str, float] | None = None,
        row_heights: dict[int, float] | None = None,
    ) -> list[list[dict[str, Any]]]:
        """
        行データを解析してリスト形式で返す（コード重複削減用ヘルパー）

        Args:
            rows: 行データのタプル
            include_cell_styles: セルのスタイル情報を含めるか
            merged_cell_map: マージセル情報
            merged_anchor_value_map: マージ範囲 -> アンカー値
            col_widths: 列幅のキャッシュ（パフォーマンス最適化用）
            row_heights: 行高さのキャッシュ（パフォーマンス最適化用）

        Returns:
            解析された行データのリスト
        """
        parsed_rows = []
        for row in rows:
            row_data = [
                self._parse_cell(
                    cell,
                    include_cell_styles,
                    merged_cell_map,
                    merged_anchor_value_map,
                    col_widths,
                    row_heights,
                )
                for cell in row
            ]
            parsed_rows.append(row_data)
        return parsed_rows

    def _serialize_value(self, value: Any) -> Any:
        """
        セル値をJSONシリアライズ可能な形式に変換

        Args:
            value: セル値

        Returns:
            JSONシリアライズ可能な値
        """
        if value is None:
            return None

        # 基本的な型（JSONシリアライズ可能）はそのまま
        if isinstance(value, (str, int, float, bool)):
            return value

        # その他の型（datetime, timedelta等）は文字列に変換
        return str(value)

    def _normalize_range_data(self, range_data: Any) -> tuple[tuple[Cell, ...], ...]:
        """
        openpyxlの範囲データを統一的なタプルのタプル形式に変換

        Args:
            range_data: sheet[range]の戻り値

        Returns:
            タプルのタプル形式の範囲データ
        """
        if isinstance(range_data, Cell):
            # 単一セルの場合
            return ((range_data,),)
        elif not range_data:
            # 空の場合
            return ()
        elif not isinstance(range_data[0], tuple):
            # 単一列/行の場合
            return (range_data,)
        else:
            # 通常の範囲の場合
            return range_data
