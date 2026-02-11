"""
Excel範囲計算ユーティリティ

セル範囲の計算・変換・検証を担当するヘルパークラス
"""

import logging

from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string

logger = logging.getLogger(__name__)


class ExcelRangeCalculator:
    """セル範囲の計算・変換・検証（全て staticmethod）"""

    @staticmethod
    def calculate_header_range(cell_range: str, frozen_rows: int) -> str | None:
        """
        セル範囲に対してfrozen_rowsに基づくヘッダー範囲を計算

        Args:
            cell_range: セル範囲（例: "A5:D10"）
                       拡張後のeffective_rangeを渡すこと（軸拡張済み）
            frozen_rows: 固定行数

        Returns:
            ヘッダー範囲（例: "A1:D2"）またはNone

        早期リターン条件:
        - frozen_rows=0: ヘッダーなし
        - start_row == 1: 既に1行目から開始（ヘッダー全体を含む）

        部分的な重なり処理:
        - frozen_rows=2, cell_range="A2:B6" の場合
          → 不足分 "A1:B1" を返して、最終的に "A1:B6" になる
        """
        # frozen_rowsが0の場合はヘッダーなし
        if frozen_rows == 0:
            return None

        # セル範囲を解析
        # "A5:D10" -> start="A5", end="D10"
        if ":" in cell_range:
            start, end = cell_range.split(":")
        else:
            # 単一セル（例: "B5"）
            start = end = cell_range

        # 開始セルの座標を解析
        start_col_letter, start_row = coordinate_from_string(start)
        end_col_letter, _ = coordinate_from_string(end)

        # 既に1行目から開始している場合は追加不要（ヘッダー全体を含む）
        if start_row == 1:
            return None

        # 部分的な重なりがある場合は、不足している上部のヘッダー行を追加
        if start_row <= frozen_rows:
            # 1行目から(start_row-1)行目までを追加
            header_range = f"{start_col_letter}1:{end_col_letter}{start_row - 1}"
            return header_range

        # ヘッダー範囲を計算: {start_col}1:{end_col}{frozen_rows}
        header_range = f"{start_col_letter}1:{end_col_letter}{frozen_rows}"
        return header_range

    @staticmethod
    def merge_ranges(range1: str, range2: str) -> str:
        """
        2つのセル範囲を結合して、最小の包含範囲を返す

        Args:
            range1: 範囲1（例: "A1:B2"）
            range2: 範囲2（例: "A4:B6"）

        Returns:
            結合された範囲（例: "A1:B6"）
        """
        # 範囲1を解析
        if ":" in range1:
            start1, end1 = range1.split(":")
        else:
            start1 = end1 = range1

        # 範囲2を解析
        if ":" in range2:
            start2, end2 = range2.split(":")
        else:
            start2 = end2 = range2

        # 座標を取得
        col1_start, row1_start = coordinate_from_string(start1)
        col1_end, row1_end = coordinate_from_string(end1)
        col2_start, row2_start = coordinate_from_string(start2)
        col2_end, row2_end = coordinate_from_string(end2)

        # 最小/最大の列を決定
        col_start_idx = min(
            column_index_from_string(col1_start), column_index_from_string(col2_start)
        )
        col_end_idx = max(
            column_index_from_string(col1_end), column_index_from_string(col2_end)
        )

        # 最小/最大の行を決定
        row_start = min(row1_start, row2_start)
        row_end = max(row1_end, row2_end)

        # 列インデックスを文字に変換
        col_start = get_column_letter(col_start_idx)
        col_end = get_column_letter(col_end_idx)

        return f"{col_start}{row_start}:{col_end}{row_end}"

    @staticmethod
    def expand_axis_range(range_str: str) -> str:
        """
        単一セル・列・行を1行目/A列まで拡張する

        Examples:
            - 単一セル: "C5" -> "C1:C5"（列方向に拡張）
            - 単一列: "Z100:Z200" -> "Z1:Z200"（上端まで拡張）
            - 単一行: "D200:Z200" -> "A200:Z200"（左端まで拡張）
            - 矩形範囲: そのまま（変更なし）

        Args:
            range_str: セル範囲文字列

        Returns:
            拡張されたセル範囲（または元の範囲）
        """
        if not range_str:
            return range_str

        raw = range_str.strip()
        if ":" not in raw:
            try:
                col, row = coordinate_from_string(raw.replace("$", ""))
                return f"{col}1:{col}{row}"
            except ValueError:
                return range_str

        start_cell, end_cell = raw.split(":", 1)
        start_cell = start_cell.replace("$", "")
        end_cell = end_cell.replace("$", "")

        start_col, start_row = coordinate_from_string(start_cell)
        end_col, end_row = coordinate_from_string(end_cell)

        # 列指定（同一列）: 逆順はそのまま（既存のrange検証で弾く）
        if start_col == end_col:
            if end_row < start_row:
                return range_str
            return f"{start_col}1:{end_col}{end_row}"

        # 行指定（同一行）: 逆順はそのまま（既存のrange検証で弾く）
        if start_row == end_row:
            if column_index_from_string(end_col) < column_index_from_string(start_col):
                return range_str
            return f"A{start_row}:{end_col}{end_row}"

        return range_str

    @staticmethod
    def calculate_range_size(range_str: str) -> tuple[int, int]:
        """
        セル範囲文字列から行数と列数を計算

        Args:
            range_str: セル範囲（例: "A1:D10" または "A1:XFD1048576"）

        Returns:
            (rows, cols)のタプル。
            エラー時は (0, 0) を返す。
        """
        try:
            if ":" not in range_str:
                # 単一セルの場合
                return (1, 1)

            start_cell, end_cell = range_str.split(":")
            start_col, start_row = coordinate_from_string(start_cell)
            end_col, end_row = coordinate_from_string(end_cell)

            start_col_idx = column_index_from_string(start_col)
            end_col_idx = column_index_from_string(end_col)

            # 逆順序の範囲を検出（セキュリティ対策）
            if end_row < start_row or end_col_idx < start_col_idx:
                raise ValueError(
                    f"無効なセル範囲: '{range_str}'。"
                    f"範囲は正しい順序で指定してください（例: 'A1:Z100'）"
                )

            rows = end_row - start_row + 1
            cols = end_col_idx - start_col_idx + 1

            return (rows, cols)
        except Exception as e:
            # 元の実装との互換性維持: エラー時は (0, 0) を返す
            logger.warning(f"Failed to calculate range size '{range_str}': {e}")
            return (0, 0)

    @staticmethod
    def normalize_column_range(cell_range: str, max_row: int) -> str:
        """
        列のみ指定された範囲（例: "J:J" / "J"）を行番号付きに正規化する

        Args:
            cell_range: セル範囲
            max_row: シートの最大行番号

        Returns:
            正規化されたセル範囲

        Raises:
            ValueError: 逆順序の列範囲を検出した場合
        """
        raw = cell_range.strip()
        if not raw:
            return cell_range

        # "J:J" のような列のみ指定
        if ":" in raw:
            start, end = raw.split(":", 1)
            start_col = start.replace("$", "")
            end_col = end.replace("$", "")
            if start_col.isalpha() and end_col.isalpha():
                start_col = start_col.upper()
                end_col = end_col.upper()
                # 逆順序の列を検出
                if column_index_from_string(end_col) < column_index_from_string(
                    start_col
                ):
                    raise ValueError(
                        f"無効なセル範囲: '{cell_range}'。"
                        f"範囲は正しい順序で指定してください（例: 'A1:Z100'）"
                    )
                return f"{start_col}1:{end_col}{max_row}"

        # "J" のような単一列指定
        col_only = raw.replace("$", "")
        if col_only.isalpha():
            col_only = col_only.upper()
            return f"{col_only}1:{col_only}{max_row}"

        return cell_range
