"""
Excelマージセル処理ユーティリティ

マージセル情報のキャッシュ構築と値伝播を担当するヘルパークラス
"""

from typing import Any

from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string


class ExcelMergedCellHandler:
    """マージセル情報の構築と管理（全て staticmethod）"""

    @staticmethod
    def build_merged_cell_cache(
        sheet,
        effective_cell_range: str | None,
        value_serializer,
    ) -> tuple[dict[str, str] | None, dict[str, Any] | None, list[dict[str, Any]]]:
        """
        マージセル情報をキャッシュして返す（パフォーマンス最適化）
        - 「今回返す予定の範囲」を先に確定し、その範囲と交差する結合だけを部分展開する
        - アンカー値は左上→無ければ結合範囲内の実在セルのみから最小(row,col)を選ぶ

        Args:
            sheet: openpyxl Worksheet
            effective_cell_range: 正規化・拡張済みのセル範囲（例: "A1:D10"）
                Noneの場合はsheet.dimensionsを使用
            value_serializer: セル値をシリアライズする関数（例: parser._serialize_value）

        Returns:
            (merged_cell_map, merged_anchor_value_map, merged_ranges)のタプル
            - merged_cell_map: セル座標 -> 結合範囲のマップ
            - merged_anchor_value_map: 結合範囲 -> アンカー値のマップ
            - merged_ranges: 結合範囲情報のリスト
        """
        merged_ranges: list[dict[str, Any]] = []

        # 今回返す予定の範囲（結合情報の部分展開に使用）
        # effective_cell_rangeがあればそれを使用、なければsheet.dimensionsを使用
        planned_range_for_merge = effective_cell_range or (
            str(sheet.dimensions) if sheet.dimensions else None
        )

        if not sheet.merged_cells.ranges or not planned_range_for_merge:
            return (None, None, [])

        # planned_range_for_merge から対象範囲の境界を計算
        if ":" in planned_range_for_merge:
            start_cell, end_cell = planned_range_for_merge.split(":", 1)
        else:
            start_cell = planned_range_for_merge
            end_cell = planned_range_for_merge

        start_cell = start_cell.replace("$", "")
        end_cell = end_cell.replace("$", "")

        start_col, start_row = coordinate_from_string(start_cell)
        end_col, end_row = coordinate_from_string(end_cell)

        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)

        target_min_row = min(start_row, end_row)
        target_max_row = max(start_row, end_row)
        target_min_col = min(start_col_idx, end_col_idx)
        target_max_col = max(start_col_idx, end_col_idx)

        merged_cell_map: dict[str, str] = {}
        merged_anchor_value_map: dict[str, Any] = {}

        for merged_range in sheet.merged_cells.ranges:
            merged_range_str = str(merged_range)
            range_start = merged_range_str.split(":")[0]

            merged_min_row = merged_range.min_row
            merged_max_row = merged_range.max_row
            merged_min_col = merged_range.min_col
            merged_max_col = merged_range.max_col

            # 返す予定の範囲と交差しない結合は無視（部分展開）
            inter_min_row = max(merged_min_row, target_min_row)
            inter_max_row = min(merged_max_row, target_max_row)
            inter_min_col = max(merged_min_col, target_min_col)
            inter_max_col = min(merged_max_col, target_max_col)
            if inter_min_row > inter_max_row or inter_min_col > inter_max_col:
                continue

            # アンカー値を決定（左上が空なら結合範囲内の実在セルだけ走査）
            anchor_coord = range_start
            anchor_value = value_serializer(sheet[range_start].value)

            if anchor_value is None:
                anchor_coord, anchor_value = (
                    ExcelMergedCellHandler._find_anchor_value_in_merge(
                        sheet,
                        merged_min_row,
                        merged_max_row,
                        merged_min_col,
                        merged_max_col,
                        value_serializer,
                    )
                )

            # セル座標 -> 結合範囲 のマップ（返す予定の範囲と交差する部分だけ展開）
            for row_idx in range(inter_min_row, inter_max_row + 1):
                for col_idx in range(inter_min_col, inter_max_col + 1):
                    coord_str = f"{get_column_letter(col_idx)}{row_idx}"
                    merged_cell_map[coord_str] = merged_range_str

            # アンカー値を保存（結合セルの値埋め用）
            merged_anchor_value_map[merged_range_str] = anchor_value

            # 結合範囲そのものを返す（結合セルがある時だけ返す）
            merged_ranges.append(
                {
                    "range": merged_range_str,
                    "anchor": {"coordinate": anchor_coord, "value": anchor_value},
                }
            )

        if not merged_ranges:
            return (None, None, [])

        return (merged_cell_map, merged_anchor_value_map, merged_ranges)

    @staticmethod
    def _find_anchor_value_in_merge(
        sheet,
        merged_min_row: int,
        merged_max_row: int,
        merged_min_col: int,
        merged_max_col: int,
        value_serializer,
    ) -> tuple[str, Any | None]:
        """
        結合セル範囲内で最初の非空値を探す（左上が空の場合）

        Args:
            sheet: openpyxl Worksheet
            merged_min_row: 結合範囲の最小行
            merged_max_row: 結合範囲の最大行
            merged_min_col: 結合範囲の最小列
            merged_max_col: 結合範囲の最大列
            value_serializer: セル値をシリアライズする関数

        Returns:
            (anchor_coord, anchor_value)のタプル
        """
        best_rc: tuple[int, int] | None = None
        best_val: Any | None = None

        # 実在セル（sheet._cells）だけから、結合範囲内の最小(row,col)の値を選ぶ
        # 互換性のため_cellsの有無をチェックしてフォールバック
        # 注意: _cellsはopenpyxlのプライベート属性のため、将来のバージョンで変更される可能性があります。
        # その場合は公開APIを使用するフォールバックロジックが動作します。
        if hasattr(sheet, "_cells"):
            # プライベート属性を使った高速版
            for (r, c), cell_obj in sheet._cells.items():
                if (
                    merged_min_row <= r <= merged_max_row
                    and merged_min_col <= c <= merged_max_col
                ):
                    cell_value = value_serializer(cell_obj.value)
                    if cell_value is not None:
                        if best_rc is None or (r, c) < best_rc:
                            best_rc = (r, c)
                            best_val = cell_value
        else:
            # 公開APIを使ったフォールバック版
            for row_idx in range(merged_min_row, merged_max_row + 1):
                for col_idx in range(merged_min_col, merged_max_col + 1):
                    coord = f"{get_column_letter(col_idx)}{row_idx}"
                    cell = sheet[coord]
                    cell_value = value_serializer(cell.value)
                    if cell_value is not None:
                        if best_rc is None or (row_idx, col_idx) < best_rc:
                            best_rc = (row_idx, col_idx)
                            best_val = cell_value

        # アンカー座標とアンカー値を返す
        if best_rc is not None:
            r, c = best_rc
            anchor_coord = f"{get_column_letter(c)}{r}"
            return (anchor_coord, best_val)
        else:
            # 全てのセルが空の場合は最初のセルを返す
            anchor_coord = f"{get_column_letter(merged_min_col)}{merged_min_row}"
            return (anchor_coord, None)
