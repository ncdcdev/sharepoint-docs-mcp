"""
ExcelPaneManagerのテスト
"""

from unittest.mock import Mock

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.views import Pane, SheetView

from src.excel import ExcelPaneManager


class TestExcelPaneManager:
    """ExcelPaneManager（固定行列処理）のテスト"""

    # get_frozen_panes のテスト

    def test_get_frozen_panes_both_rows_and_cols(self):
        """行と列の両方が固定されている場合"""
        wb = Workbook()
        ws = wb.active

        # freeze_panes="B3" (frozen_rows=2, frozen_cols=1)
        ws.freeze_panes = "B3"

        frozen_rows, frozen_cols = ExcelPaneManager.get_frozen_panes(ws)
        assert frozen_rows == 2
        assert frozen_cols == 1

    def test_get_frozen_panes_rows_only(self):
        """行のみが固定されている場合"""
        wb = Workbook()
        ws = wb.active

        # freeze_panes="A3" (frozen_rows=2, frozen_cols=0)
        ws.freeze_panes = "A3"

        frozen_rows, frozen_cols = ExcelPaneManager.get_frozen_panes(ws)
        assert frozen_rows == 2
        assert frozen_cols == 0

    def test_get_frozen_panes_cols_only(self):
        """列のみが固定されている場合"""
        wb = Workbook()
        ws = wb.active

        # freeze_panes="B1" (frozen_rows=0, frozen_cols=1)
        ws.freeze_panes = "B1"

        frozen_rows, frozen_cols = ExcelPaneManager.get_frozen_panes(ws)
        assert frozen_rows == 0
        assert frozen_cols == 1

    def test_get_frozen_panes_none(self):
        """freeze_panesが設定されていない場合"""
        wb = Workbook()
        ws = wb.active

        frozen_rows, frozen_cols = ExcelPaneManager.get_frozen_panes(ws)
        assert frozen_rows == 0
        assert frozen_cols == 0

    def test_get_frozen_panes_split_state_ignored(self):
        """split state（state="split"）は無視されること"""
        wb = Workbook()
        ws = wb.active

        # split paneを設定（frozenではなくsplit）
        ws.sheet_view.pane = Pane(ySplit=3, xSplit=2, state="split")

        frozen_rows, frozen_cols = ExcelPaneManager.get_frozen_panes(ws)
        # split stateは無視される
        assert frozen_rows == 0
        assert frozen_cols == 0

    def test_get_frozen_panes_with_scroll_position(self):
        """スクロール位置（topLeftCell）に影響されないこと"""
        wb = Workbook()
        ws = wb.active

        # 3行固定を設定
        ws.freeze_panes = "A4"

        # スクロール位置を変更（pane.topLeftCellを直接操作）
        pane = ws.sheet_view.pane
        pane.topLeftCell = "A450"

        frozen_rows, frozen_cols = ExcelPaneManager.get_frozen_panes(ws)
        # pane.ySplit=3なので、frozen_rowsは3であるべき（449ではない）
        assert frozen_rows == 3
        assert frozen_cols == 0

    def test_get_frozen_panes_with_large_values(self):
        """大きな固定行数・列数でも正しく取得できること"""
        wb = Workbook()
        ws = wb.active

        # 異常に大きなfrozen_rowsを設定
        pane = Pane()
        pane.ySplit = 200
        pane.xSplit = 50
        pane.topLeftCell = "AY201"
        pane.state = "frozen"

        sheet_view = SheetView(pane=pane)
        ws.views.sheetView[0] = sheet_view

        frozen_rows, frozen_cols = ExcelPaneManager.get_frozen_panes(ws)
        assert frozen_rows == 200
        assert frozen_cols == 50

    def test_get_frozen_panes_exception_handling(self):
        """例外が発生した場合は(0, 0)を返すこと"""
        # sheet.sheet_viewがNoneのモックを作成
        mock_sheet = Mock()
        mock_sheet.sheet_view = None

        frozen_rows, frozen_cols = ExcelPaneManager.get_frozen_panes(mock_sheet)
        assert frozen_rows == 0
        assert frozen_cols == 0

    # format_freeze_panes のテスト

    def test_format_freeze_panes_both(self):
        """行と列の両方が固定されている場合"""
        result = ExcelPaneManager.format_freeze_panes(2, 1)
        assert result == "B3"

    def test_format_freeze_panes_rows_only(self):
        """行のみが固定されている場合"""
        result = ExcelPaneManager.format_freeze_panes(2, 0)
        assert result == "A3"

    def test_format_freeze_panes_cols_only(self):
        """列のみが固定されている場合"""
        result = ExcelPaneManager.format_freeze_panes(0, 1)
        assert result == "B1"

    def test_format_freeze_panes_zero(self):
        """固定なしの場合"""
        result = ExcelPaneManager.format_freeze_panes(0, 0)
        assert result == "A1"

    def test_format_freeze_panes_large_values(self):
        """大きな値でも正しくフォーマットされること"""
        result = ExcelPaneManager.format_freeze_panes(100, 25)
        # 26列目（Z）の101行目
        assert result == "Z101"

    # validate_frozen_rows のテスト

    def test_validate_frozen_rows_within_limit(self):
        """上限以内の場合はTrueと元の値が返ること"""
        is_valid, validated = ExcelPaneManager.validate_frozen_rows(50, 100)
        assert is_valid is True
        assert validated == 50

    def test_validate_frozen_rows_at_limit(self):
        """上限と同じ値の場合はTrueと元の値が返ること"""
        is_valid, validated = ExcelPaneManager.validate_frozen_rows(100, 100)
        assert is_valid is True
        assert validated == 100

    def test_validate_frozen_rows_exceeds_limit(self):
        """上限を超える場合はFalseと0が返ること"""
        is_valid, validated = ExcelPaneManager.validate_frozen_rows(150, 100)
        assert is_valid is False
        assert validated == 0

    def test_validate_frozen_rows_zero(self):
        """0の場合はTrueと0が返ること"""
        is_valid, validated = ExcelPaneManager.validate_frozen_rows(0, 100)
        assert is_valid is True
        assert validated == 0

    def test_validate_frozen_rows_negative_value(self):
        """負の値の場合はTrueと0が返ること（防御的処理）"""
        is_valid, validated = ExcelPaneManager.validate_frozen_rows(-1, 100)
        assert is_valid is True
        assert validated == 0
