import datetime
import json
from io import BytesIO
from unittest.mock import Mock, patch
from zipfile import BadZipFile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.views import Pane, SheetView

from src.sharepoint_excel import SharePointExcelParser


class TestSharePointExcelParser:
    """SharePoint Excel解析のテスト"""

    def setup_method(self):
        """テストメソッド実行前のセットアップ"""
        self.mock_download_client = Mock()

    def _create_test_excel(self) -> bytes:
        """テスト用のシンプルなExcelファイルを作成"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Name"
        ws["B1"] = "Age"
        ws["A2"] = "John"
        ws["B2"] = 25

        # BytesIOに保存
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        return excel_bytes.getvalue()

    def _create_formatted_excel(self) -> bytes:
        """書式を含むテスト用Excelファイルを作成"""
        wb = Workbook()
        ws = wb.active
        ws.title = "FormattedSheet"

        # ヘッダー行に書式を設定
        ws["A1"] = "Name"
        ws["A1"].font = Font(name="Arial", size=12, bold=True, color="FF0000")
        ws["A1"].fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        ws["B1"] = "Value"
        ws["B1"].font = Font(name="Arial", size=12, bold=True)

        # データ行
        ws["A2"] = "Item1"
        ws["B2"] = 100

        # BytesIOに保存
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        return excel_bytes.getvalue()

    def _create_multi_sheet_excel(self) -> bytes:
        """複数シートを含むテスト用Excelファイルを作成"""
        wb = Workbook()

        # 最初のシート
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"] = "Data1"

        # 2つ目のシート
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "Data2"

        # BytesIOに保存
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        return excel_bytes.getvalue()

    def _create_merged_cells_excel(self) -> bytes:
        """結合セルを含むテスト用Excelファイルを作成"""
        wb = Workbook()
        ws = wb.active
        ws.title = "MergedSheet"

        # セルを結合
        ws.merge_cells("A1:B1")
        ws["A1"] = "Merged Header"

        ws["A2"] = "Data1"
        ws["B2"] = "Data2"

        # BytesIOに保存
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        return excel_bytes.getvalue()

    def _create_frozen_panes_excel(self, freeze_panes: str) -> bytes:
        """固定行・列を含むテスト用Excelファイルを作成"""
        wb = Workbook()
        ws = wb.active
        ws.title = "FrozenSheet"

        # ヘッダー行を作成
        ws["A1"] = "Header1"
        ws["B1"] = "Header2"
        ws["C1"] = "Header3"
        ws["D1"] = "Header4"

        # データ行を作成
        for row in range(2, 11):
            for col in range(1, 5):
                ws.cell(row=row, column=col, value=f"Data{row - 1}_{col}")

        # freeze_panesを設定
        ws.freeze_panes = freeze_panes

        # BytesIOに保存
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        return excel_bytes.getvalue()

    def test_parse_simple_excel(self):
        """シンプルなExcelファイルの解析テスト（デフォルト：最小限の情報）"""
        excel_bytes = self._create_test_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        # 解析
        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.parse_to_json("/test/file.xlsx")

        # 検証
        result = json.loads(result_json)
        assert result["file_path"] == "/test/file.xlsx"
        assert len(result["sheets"]) == 1
        assert result["sheets"][0]["name"] == "Sheet1"
        assert len(result["sheets"][0]["rows"]) == 2

        # デフォルトでは value と coordinate のみ
        cell = result["sheets"][0]["rows"][0][0]
        assert cell["value"] == "Name"
        assert cell["coordinate"] == "A1"
        assert "data_type" not in cell
        assert "fill" not in cell
        assert "width" not in cell

        cell2 = result["sheets"][0]["rows"][1][1]
        assert cell2["value"] == 25
        assert cell2["coordinate"] == "B2"

    def test_parse_multiple_sheets(self):
        """複数シートのExcelファイルの解析テスト"""
        excel_bytes = self._create_multi_sheet_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.parse_to_json("/test/multi.xlsx")

        result = json.loads(result_json)
        assert len(result["sheets"]) == 2
        assert result["sheets"][0]["name"] == "Sheet1"
        assert result["sheets"][1]["name"] == "Sheet2"
        assert result["sheets"][0]["rows"][0][0]["value"] == "Data1"
        assert result["sheets"][1]["rows"][0][0]["value"] == "Data2"

    def test_parse_merged_cells(self):
        """結合セルの解析テスト（結合情報は常に含まれる）"""
        excel_bytes = self._create_merged_cells_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.parse_to_json("/test/merged.xlsx")

        result = json.loads(result_json)
        assert result["sheets"][0]["name"] == "MergedSheet"

        # 結合セルの情報を確認
        merged_cell = result["sheets"][0]["rows"][0][0]
        assert merged_cell["value"] == "Merged Header"
        assert "merged" in merged_cell
        assert merged_cell["merged"]["range"] == "A1:B1"
        assert merged_cell["merged"]["is_top_left"] is True

    def test_download_error_handling(self):
        """ダウンロードエラーのハンドリングテスト"""
        self.mock_download_client.download_file.side_effect = Exception(
            "Download failed"
        )

        parser = SharePointExcelParser(self.mock_download_client)
        with pytest.raises(Exception) as exc_info:
            parser.parse_to_json("/test/file.xlsx")

        assert "Download failed" in str(exc_info.value)

    def test_invalid_excel_file(self):
        """無効なExcelファイルの処理テスト"""
        # 無効なバイトデータを返す
        self.mock_download_client.download_file.return_value = b"invalid excel data"

        parser = SharePointExcelParser(self.mock_download_client)
        with pytest.raises((BadZipFile, InvalidFileException)):
            parser.parse_to_json("/test/invalid.xlsx")

    def test_empty_excel_file(self):
        """空のExcelファイルの解析テスト"""
        wb = Workbook()
        ws = wb.active
        ws.title = "EmptySheet"
        # データを追加しない

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        self.mock_download_client.download_file.return_value = excel_bytes.getvalue()

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.parse_to_json("/test/empty.xlsx")

        result = json.loads(result_json)
        assert result["sheets"][0]["name"] == "EmptySheet"
        # openpyxlは空のシートでも最低1つのセル（A1）を持つ
        assert result["sheets"][0]["dimensions"] is not None
        # 空のシートでも行データが取得できる
        rows = result["sheets"][0]["rows"]
        assert isinstance(rows, list)

    def test_color_to_hex_rgb(self):
        """RGB色の16進数変換テスト"""
        excel_bytes = self._create_formatted_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        # _color_to_hex の単体動作を確認（include_formattingの有無とは無関係）
        parser = SharePointExcelParser(self.mock_download_client)
        wb = load_workbook(BytesIO(excel_bytes))
        cell = wb.active["A1"]
        hex_color = parser._color_to_hex(cell.fill.fgColor)
        if hex_color:
            assert hex_color.startswith("#")

    def test_parse_with_formulas(self):
        """数式を含むExcelファイルの解析テスト"""
        wb = Workbook()
        ws = wb.active
        ws.title = "FormulaSheet"

        ws["A1"] = 10
        ws["A2"] = 20
        ws["A3"] = "=A1+A2"  # 数式

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        self.mock_download_client.download_file.return_value = excel_bytes.getvalue()

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.parse_to_json("/test/formula.xlsx")

        result = json.loads(result_json)
        # 数式セルの値を確認（data_only=Falseなので数式文字列が入る）
        formula_cell = result["sheets"][0]["rows"][2][0]
        assert formula_cell["value"] == "=A1+A2"

    def test_default_response_is_minimal(self):
        """デフォルトレスポンスが最小限であることのテスト"""
        excel_bytes = self._create_formatted_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.parse_to_json("/test/formatted.xlsx")

        result = json.loads(result_json)
        cell = result["sheets"][0]["rows"][0][0]

        # デフォルトでは value と coordinate のみ
        assert "value" in cell
        assert "coordinate" in cell
        assert "data_type" not in cell
        assert "fill" not in cell
        assert "font" not in cell
        assert "alignment" not in cell
        assert "merged" not in cell
        assert "width" not in cell
        assert "height" not in cell

    def test_datetime_serialization(self):
        """datetime型の値が正しくシリアライズされることのテスト"""
        wb = Workbook()
        ws = wb.active
        ws.title = "DateTimeSheet"

        # 各種datetime型の値を設定
        ws["A1"] = datetime.datetime(2024, 1, 15, 14, 30, 45)
        ws["A2"] = datetime.date(2024, 1, 15)
        ws["A3"] = datetime.time(14, 30, 45)
        ws["A4"] = datetime.timedelta(days=1, hours=2, minutes=30)

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        self.mock_download_client.download_file.return_value = excel_bytes.getvalue()

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.parse_to_json("/test/datetime.xlsx")

        # JSONパースが成功することを確認（datetime型が適切にシリアライズされている）
        result = json.loads(result_json)
        rows = result["sheets"][0]["rows"]

        # 文字列に変換されていることを確認
        assert rows[0][0]["value"] == "2024-01-15 14:30:45"
        # openpyxlはdateをdatetimeとして読み込む（時刻部分は00:00:00）
        assert rows[1][0]["value"] == "2024-01-15 00:00:00"
        assert rows[2][0]["value"] == "14:30:45"
        # timedeltaは文字列表現に変換される
        assert rows[3][0]["value"] == "1 day, 2:30:00"

    def test_search_cells_basic(self):
        """セル検索の基本テスト"""
        excel_bytes = self._create_test_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.search_cells("/test/file.xlsx", "John")

        result = json.loads(result_json)
        assert result["file_path"] == "/test/file.xlsx"
        assert result["mode"] == "search"
        assert result["query"] == "John"
        assert result["match_count"] == 1
        assert len(result["matches"]) == 1
        assert result["matches"][0]["sheet"] == "Sheet1"
        assert result["matches"][0]["coordinate"] == "A2"
        assert result["matches"][0]["value"] == "John"

    def test_search_cells_multiple_matches(self):
        """複数マッチする検索のテスト"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "売上報告"
        ws["A2"] = "月間売上"
        ws["B2"] = 1000
        ws["A3"] = "売上合計"
        ws["B3"] = "=SUM(B1:B2)"

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        self.mock_download_client.download_file.return_value = excel_bytes.getvalue()

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.search_cells("/test/file.xlsx", "売上")

        result = json.loads(result_json)
        assert result["match_count"] == 3
        assert len(result["matches"]) == 3
        coordinates = [m["coordinate"] for m in result["matches"]]
        assert "A1" in coordinates
        assert "A2" in coordinates
        assert "A3" in coordinates

    def test_search_cells_no_match(self):
        """マッチしない検索のテスト"""
        excel_bytes = self._create_test_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.search_cells("/test/file.xlsx", "NotFound")

        result = json.loads(result_json)
        assert result["match_count"] == 0
        assert result["matches"] == []

    def test_search_cells_multiple_sheets(self):
        """複数シートにまたがる検索のテスト"""
        excel_bytes = self._create_multi_sheet_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.search_cells("/test/file.xlsx", "Data")

        result = json.loads(result_json)
        assert result["match_count"] == 2
        assert len(result["matches"]) == 2
        sheets = [m["sheet"] for m in result["matches"]]
        assert "Sheet1" in sheets
        assert "Sheet2" in sheets

    def _create_test_data_excel(self) -> bytes:
        """テスト用のデータExcelファイルを作成（検索テスト用）"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # ヘッダー行
        ws["A1"] = "ID"
        ws["B1"] = "名前"
        ws["C1"] = "金額"
        ws["D1"] = "備考"
        # データ行
        ws["A2"] = 1
        ws["B2"] = "商品A"
        ws["C2"] = 1000
        ws["D2"] = "在庫あり"
        ws["A3"] = 2
        ws["B3"] = "商品B"
        ws["C3"] = 2000
        ws["D3"] = "売上好調"
        ws["A4"] = 3
        ws["B4"] = "商品C"
        ws["C4"] = 1500
        ws["D4"] = "在庫わずか"

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        return excel_bytes.getvalue()

    def _create_search_test_excel(self) -> bytes:
        """複数キーワード検索テスト用のExcelファイルを作成（AND検索用）"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "2024年度 予算 報告書"
        ws["A2"] = "売上予測"
        ws["A3"] = "経費明細"
        ws["A4"] = "利益 計算 シート"
        ws["A5"] = "予算"

        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "予算案 データ"
        ws2["A2"] = "データ分析"

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        return excel_bytes.getvalue()

    def test_search_with_surrounding_cells_disabled(self):
        """デフォルト動作（include_surrounding_cells=False）の確認"""
        excel_bytes = self._create_test_data_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.search_cells(
            "/test/file.xlsx", "商品B", include_surrounding_cells=False
        )

        result = json.loads(result_json)
        assert result["match_count"] == 1
        assert len(result["matches"]) == 1

        match = result["matches"][0]
        assert match["coordinate"] == "B3"
        assert match["value"] == "商品B"
        # row_dataは含まれない
        assert "row_data" not in match

    def test_search_with_surrounding_cells_enabled(self):
        """include_surrounding_cells=Trueで行データ取得"""
        excel_bytes = self._create_test_data_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.search_cells(
            "/test/file.xlsx", "商品B", include_surrounding_cells=True
        )

        result = json.loads(result_json)
        assert result["match_count"] == 1
        match = result["matches"][0]

        # row_dataが含まれる
        assert "row_data" in match
        row_data = match["row_data"]
        assert len(row_data) == 4  # A3, B3, C3, D3

        # 各セルのデータを確認
        assert row_data[0]["coordinate"] == "A3"
        assert row_data[0]["value"] == 2
        assert row_data[1]["coordinate"] == "B3"
        assert row_data[1]["value"] == "商品B"
        assert row_data[2]["coordinate"] == "C3"
        assert row_data[2]["value"] == 2000
        assert row_data[3]["coordinate"] == "D3"
        assert row_data[3]["value"] == "売上好調"

    def test_search_with_surrounding_cells_multiple_matches(self):
        """複数マッチ時の行データ取得"""
        excel_bytes = self._create_test_data_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.search_cells(
            "/test/file.xlsx", "在庫", include_surrounding_cells=True
        )

        result = json.loads(result_json)
        assert result["match_count"] == 2
        assert len(result["matches"]) == 2

        # 1つ目のマッチ（D2: "在庫あり"）
        match1 = result["matches"][0]
        assert match1["coordinate"] == "D2"
        assert match1["value"] == "在庫あり"
        assert "row_data" in match1
        assert len(match1["row_data"]) == 4
        assert match1["row_data"][1]["value"] == "商品A"

        # 2つ目のマッチ（D4: "在庫わずか"）
        match2 = result["matches"][1]
        assert match2["coordinate"] == "D4"
        assert match2["value"] == "在庫わずか"
        assert "row_data" in match2
        assert len(match2["row_data"]) == 4
        assert match2["row_data"][1]["value"] == "商品C"

    def test_search_with_surrounding_cells_empty_cells(self):
        """空セル（None）も含まれること確認"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "データ"
        ws["B1"] = None  # 空セル
        ws["C1"] = "情報"

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        self.mock_download_client.download_file.return_value = excel_bytes.getvalue()

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.search_cells(
            "/test/file.xlsx", "データ", include_surrounding_cells=True
        )

        result = json.loads(result_json)
        assert result["match_count"] == 1
        match = result["matches"][0]

        # row_dataに空セルも含まれる
        assert "row_data" in match
        row_data = match["row_data"]
        assert len(row_data) == 3
        assert row_data[0]["value"] == "データ"
        assert row_data[1]["value"] is None  # 空セル
        assert row_data[2]["value"] == "情報"

    def test_search_single_keyword_backward_compatible(self):
        """単一キーワード（後方互換性）"""
        excel_bytes = self._create_search_test_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.search_cells("/test/file.xlsx", "予算")

        result = json.loads(result_json)
        assert result["match_count"] == 3
        # "2024年度 予算 報告書", "予算", "予算案 データ"
        values = [m["value"] for m in result["matches"]]
        assert "2024年度 予算 報告書" in values
        assert "予算" in values
        assert "予算案 データ" in values

    def test_search_multiple_keywords_space_separated(self):
        """スペース区切りAND検索"""
        excel_bytes = self._create_search_test_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.search_cells("/test/file.xlsx", "予算 報告")

        result = json.loads(result_json)
        # "2024年度 予算 報告書"のみがマッチ（両方のキーワードを含む）
        assert result["match_count"] == 1
        values = [m["value"] for m in result["matches"]]
        assert "2024年度 予算 報告書" in values

    def test_search_multiple_keywords_with_extra_spaces(self):
        """前後の余分なスペースの処理"""
        excel_bytes = self._create_search_test_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        # 余分なスペースを含むキーワード指定
        result_json = parser.search_cells("/test/file.xlsx", "  予算  報告  ")

        result = json.loads(result_json)
        # スペースがトリムされて正しくマッチ
        assert result["match_count"] == 1
        values = [m["value"] for m in result["matches"]]
        assert "2024年度 予算 報告書" in values

    def test_search_multiple_keywords_no_match(self):
        """全キーワードを含むセルがない場合"""
        excel_bytes = self._create_search_test_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        # "予算"を含むが"在庫"を含まないため、マッチなし
        result_json = parser.search_cells("/test/file.xlsx", "予算 在庫")

        result = json.loads(result_json)
        assert result["match_count"] == 0
        assert result["matches"] == []

    def test_search_multiple_keywords_across_sheets(self):
        """複数シートにまたがるAND検索"""
        excel_bytes = self._create_search_test_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.search_cells("/test/file.xlsx", "予算 データ")

        result = json.loads(result_json)
        # "予算案 データ"(Sheet2)のみがマッチ
        assert result["match_count"] == 1
        assert result["matches"][0]["sheet"] == "Sheet2"
        assert result["matches"][0]["value"] == "予算案 データ"

    def test_search_multiple_keywords_with_surrounding_cells(self):
        """AND検索とinclude_surrounding_cellsの組み合わせ"""
        excel_bytes = self._create_search_test_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.search_cells(
            "/test/file.xlsx", "利益 計算", include_surrounding_cells=True
        )

        result = json.loads(result_json)
        assert result["match_count"] == 1

        # マッチにrow_dataが含まれる
        match = result["matches"][0]
        assert "row_data" in match
        assert len(match["row_data"]) > 0
        assert match["value"] == "利益 計算 シート"

    def test_parse_specific_sheet(self):
        """特定シートのみ取得するテスト"""
        excel_bytes = self._create_multi_sheet_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.parse_to_json("/test/file.xlsx", sheet_name="Sheet2")

        result = json.loads(result_json)
        assert len(result["sheets"]) == 1
        assert result["sheets"][0]["name"] == "Sheet2"
        assert result["sheets"][0]["rows"][0][0]["value"] == "Data2"

    def test_parse_nonexistent_sheet(self):
        """存在しないシート名を指定した場合の解決情報テスト"""
        excel_bytes = self._create_test_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.parse_to_json("/test/file.xlsx", sheet_name="NonExistent")

        result = json.loads(result_json)
        assert result["requested_sheet"] == "NonExistent"
        assert result["sheets"] == []
        assert result["sheet_resolution"]["status"] == "not_found"
        assert result["sheet_resolution"]["requested"] == "NonExistent"
        assert result["sheet_resolution"]["resolved"] is None
        assert result["available_sheets"] == ["Sheet1"]
        assert result["warning"] == "requested sheet_name was not found or ambiguous"

    def test_resolve_sheet_name_exact_match(self):
        """_resolve_sheet_name: 完全一致のテスト"""
        parser = SharePointExcelParser(self.mock_download_client)
        sheetnames = ["Sheet1", "Sheet2", "Data"]

        resolved, candidates = parser._resolve_sheet_name(sheetnames, "Sheet1")
        assert resolved == "Sheet1"
        assert candidates == []

    def test_resolve_sheet_name_case_insensitive(self):
        """_resolve_sheet_name: 大小文字の違いで解決されるテスト"""
        parser = SharePointExcelParser(self.mock_download_client)
        sheetnames = ["MySheet", "OtherSheet"]

        # 小文字で検索 -> 大文字のシートに解決
        resolved, candidates = parser._resolve_sheet_name(sheetnames, "mysheet")
        assert resolved == "MySheet"
        assert candidates == []

        # 大文字で検索 -> 小文字混在のシートに解決
        resolved, candidates = parser._resolve_sheet_name(sheetnames, "MYSHEET")
        assert resolved == "MySheet"
        assert candidates == []

    def test_resolve_sheet_name_with_whitespace(self):
        """_resolve_sheet_name: 前後の空白で解決されるテスト"""
        parser = SharePointExcelParser(self.mock_download_client)
        sheetnames = ["Data", "Summary"]

        # 前後に空白があっても解決
        resolved, candidates = parser._resolve_sheet_name(sheetnames, "  Data  ")
        assert resolved == "Data"
        assert candidates == []

        # 空白と大小文字の組み合わせ
        resolved, candidates = parser._resolve_sheet_name(sheetnames, "  summary  ")
        assert resolved == "Summary"
        assert candidates == []

    def test_resolve_sheet_name_ambiguous_normalization(self):
        """_resolve_sheet_name: 正規化が衝突して曖昧になるケースのテスト"""
        parser = SharePointExcelParser(self.mock_download_client)
        # 正規化すると同じになる複数のシート名
        sheetnames = ["MySheet", "mysheet", "MYSHEET"]

        # 完全一致がない場合、複数候補が返る
        # strip() + casefold() で正規化（スペース除去はしない）
        resolved, candidates = parser._resolve_sheet_name(sheetnames, " mysheet ")
        assert resolved is None
        assert set(candidates) == {"MySheet", "mysheet", "MYSHEET"}

    def test_resolve_sheet_name_fuzzy_suggestions(self):
        """_resolve_sheet_name: 類似名候補を返すテスト"""
        parser = SharePointExcelParser(self.mock_download_client)
        sheetnames = ["DataSheet", "DataTable", "Summary"]

        # 類似名候補を取得
        resolved, suggestions = parser._resolve_sheet_name(sheetnames, "DataSheat")
        assert resolved is None
        assert "DataSheet" in suggestions
        assert len(suggestions) > 0

    def test_resolve_sheet_name_not_found(self):
        """_resolve_sheet_name: 見つからない場合のテスト"""
        parser = SharePointExcelParser(self.mock_download_client)
        sheetnames = ["Sheet1", "Sheet2"]

        resolved, suggestions = parser._resolve_sheet_name(
            sheetnames, "CompletelyDifferent"
        )
        assert resolved is None
        # 類似度が低すぎる場合は候補なし
        assert len(suggestions) == 0

    def test_parse_cell_range(self):
        """セル範囲指定のテスト"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # 5x5のデータを作成
        for row in range(1, 6):
            for col in range(1, 6):
                ws.cell(row=row, column=col, value=f"R{row}C{col}")

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        self.mock_download_client.download_file.return_value = excel_bytes.getvalue()

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.parse_to_json(
            "/test/file.xlsx", sheet_name="Sheet1", cell_range="B2:D4"
        )

        result = json.loads(result_json)
        assert result["sheets"][0]["requested_range"] == "B2:D4"
        # 3行3列のデータが取得される
        assert len(result["sheets"][0]["rows"]) == 3
        assert len(result["sheets"][0]["rows"][0]) == 3
        # 最初のセルはB2（R2C2）
        assert result["sheets"][0]["rows"][0][0]["value"] == "R2C2"
        assert result["sheets"][0]["rows"][0][0]["coordinate"] == "B2"
        # 最後のセルはD4（R4C4）
        assert result["sheets"][0]["rows"][2][2]["value"] == "R4C4"
        assert result["sheets"][0]["rows"][2][2]["coordinate"] == "D4"

    def test_parse_single_cell_range(self):
        """単一セル範囲指定のテスト"""
        excel_bytes = self._create_test_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.parse_to_json(
            "/test/file.xlsx", sheet_name="Sheet1", cell_range="A1"
        )

        result = json.loads(result_json)
        assert result["sheets"][0]["requested_range"] == "A1"
        assert len(result["sheets"][0]["rows"]) == 1
        assert len(result["sheets"][0]["rows"][0]) == 1
        assert result["sheets"][0]["rows"][0][0]["value"] == "Name"

    def test_parse_sheet_and_range_combined(self):
        """シート名と範囲を組み合わせたテスト"""
        excel_bytes = self._create_multi_sheet_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.parse_to_json(
            "/test/file.xlsx", sheet_name="Sheet1", cell_range="A1"
        )

        result = json.loads(result_json)
        assert len(result["sheets"]) == 1
        assert result["sheets"][0]["name"] == "Sheet1"
        assert result["sheets"][0]["rows"][0][0]["value"] == "Data1"

    def test_parse_with_freeze_panes_both(self):
        """freeze_panes="B2"（行と列の両方固定）のテスト"""
        excel_bytes = self._create_frozen_panes_excel("B2")
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.parse_to_json("/test/frozen.xlsx")

        result = json.loads(result_json)
        sheet = result["sheets"][0]

        # freeze_panes情報を確認
        assert sheet["freeze_panes"] == "B2"
        assert sheet["frozen_rows"] == 1
        assert sheet["frozen_cols"] == 1

        rows = sheet["rows"]
        assert len(rows) == 10
        assert rows[0][0]["value"] == "Header1"
        assert rows[0][1]["value"] == "Header2"
        assert rows[1][0]["value"] == "Data1_1"
        assert rows[1][1]["value"] == "Data1_2"

    def test_parse_with_freeze_panes_rows_only(self):
        """freeze_panes="A2"（行のみ固定）のテスト"""
        excel_bytes = self._create_frozen_panes_excel("A2")
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.parse_to_json("/test/frozen.xlsx")

        result = json.loads(result_json)
        sheet = result["sheets"][0]

        # freeze_panes情報を確認
        assert sheet["freeze_panes"] == "A2"
        assert sheet["frozen_rows"] == 1
        assert sheet["frozen_cols"] == 0

        rows = sheet["rows"]
        assert len(rows) == 10
        assert rows[0][0]["value"] == "Header1"
        assert rows[1][0]["value"] == "Data1_1"

    def test_parse_with_no_freeze_panes(self):
        """freeze_panes=Noneのテスト"""
        excel_bytes = self._create_test_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.parse_to_json("/test/file.xlsx")

        result = json.loads(result_json)
        sheet = result["sheets"][0]

        # freeze_panes情報を確認
        assert "freeze_panes" not in sheet
        assert sheet["frozen_rows"] == 0
        assert sheet["frozen_cols"] == 0

        rows = sheet["rows"]
        assert len(rows) == 2
        assert rows[0][0]["value"] == "Name"

    def test_parse_range_with_overlapping_headers(self):
        """cell_range内にヘッダーが含まれる場合のテスト"""
        excel_bytes = self._create_frozen_panes_excel("B2")
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        # include_frozen_rows=Falseを指定（既存の動作を維持）
        result_json = parser.parse_to_json(
            "/test/frozen.xlsx", cell_range="A1:D5", include_frozen_rows=False
        )

        result = json.loads(result_json)
        sheet = result["sheets"][0]

        rows = sheet["rows"]
        assert len(rows) == 5
        assert rows[0][0]["value"] == "Header1"
        assert rows[1][0]["value"] == "Data1_1"
        assert rows[4][0]["value"] == "Data4_1"

    def test_parse_range_with_non_overlapping_headers(self):
        """ヘッダーがcell_range外にある場合のテスト"""
        excel_bytes = self._create_frozen_panes_excel("B2")
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        # include_frozen_rows=Falseを指定（既存の動作を維持）
        result_json = parser.parse_to_json(
            "/test/frozen.xlsx", cell_range="A5:D10", include_frozen_rows=False
        )

        result = json.loads(result_json)
        sheet = result["sheets"][0]

        # requested_rangeは元のまま
        assert sheet["requested_range"] == "A5:D10"

        rows = sheet["rows"]
        assert len(rows) == 6
        assert rows[0][0]["value"] == "Data4_1"
        assert rows[0][0]["coordinate"] == "A5"

    def test_freeze_panes_scrolled_position_does_not_affect_frozen_rows(self):
        """スクロール後に保存されたファイルでfrozen_rowsが正しく取得されるテスト

        Excel上で3行固定して行450付近にスクロールして保存すると、
        pane.topLeftCell="A450"になるが、pane.ySplit=3は不変。
        旧実装ではsheet.freeze_panes（=topLeftCell）を解析していたため
        frozen_rows=449と誤判定していたバグを検証する。
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "ScrolledSheet"

        # ヘッダー3行 + データ行
        for row in range(1, 11):
            ws.cell(row=row, column=1, value=f"Row{row}")

        # 3行固定を設定
        ws.freeze_panes = "A4"

        # スクロール位置を変更（pane.topLeftCellを直接操作）
        pane = ws.sheet_view.pane
        pane.topLeftCell = "A450"

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        self.mock_download_client.download_file.return_value = excel_bytes.getvalue()

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.parse_to_json("/test/scrolled.xlsx")

        result = json.loads(result_json)
        sheet = result["sheets"][0]

        # pane.ySplit=3なので、frozen_rowsは3であるべき（449ではない）
        assert sheet["frozen_rows"] == 3
        assert sheet["frozen_cols"] == 0
        assert sheet["freeze_panes"] == "A4"

        rows = sheet["rows"]
        assert len(rows) == 10
        assert rows[0][0]["value"] == "Row1"
        assert rows[2][0]["value"] == "Row3"

    def test_split_pane_is_ignored(self):
        """split pane（state="split"）は固定行として認識されないテスト"""
        wb = Workbook()
        ws = wb.active
        ws.title = "SplitSheet"

        ws["A1"] = "Header"
        ws["A2"] = "Data"

        # split paneを設定（frozenではなくsplit）
        ws.sheet_view.pane = Pane(ySplit=3, xSplit=0, state="split")

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        self.mock_download_client.download_file.return_value = excel_bytes.getvalue()

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.parse_to_json("/test/split.xlsx")

        result = json.loads(result_json)
        sheet = result["sheets"][0]

        # split paneは固定行として認識されない
        assert sheet["frozen_rows"] == 0
        assert sheet["frozen_cols"] == 0
        assert "freeze_panes" not in sheet

    def test_normalize_column_range_single_column(self):
        """単一列指定（"J"）の正規化テスト"""
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # max_row を設定するためにデータを追加
        for i in range(1, 101):
            ws[f"A{i}"] = f"Data{i}"

        parser = SharePointExcelParser(self.mock_download_client)

        # "J" -> "J1:J100"
        normalized = parser._normalize_column_range("J", ws)
        assert normalized == "J1:J100"

        # "$J" も同様
        normalized = parser._normalize_column_range("$J", ws)
        assert normalized == "J1:J100"

        # 小文字も大文字に変換
        normalized = parser._normalize_column_range("j", ws)
        assert normalized == "J1:J100"

    def test_normalize_column_range_column_range(self):
        """列範囲指定（"J:K"）の正規化テスト"""
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # max_row を設定
        for i in range(1, 51):
            ws[f"A{i}"] = f"Data{i}"

        parser = SharePointExcelParser(self.mock_download_client)

        # "J:K" -> "J1:K50"
        normalized = parser._normalize_column_range("J:K", ws)
        assert normalized == "J1:K50"

        # "$J:$K" も同様
        normalized = parser._normalize_column_range("$J:$K", ws)
        assert normalized == "J1:K50"

        # 小文字も大文字に変換
        normalized = parser._normalize_column_range("j:k", ws)
        assert normalized == "J1:K50"

    def test_normalize_column_range_empty_sheet(self):
        """空シートでの列範囲正規化テスト（max_row=1になる）"""
        wb = Workbook()
        ws = wb.active
        ws.title = "EmptySheet"

        parser = SharePointExcelParser(self.mock_download_client)

        # 空シートの場合、max_rowは1になる
        normalized = parser._normalize_column_range("A", ws)
        assert normalized == "A1:A1"

        normalized = parser._normalize_column_range("A:C", ws)
        assert normalized == "A1:C1"

    def test_normalize_column_range_reverse_order(self):
        """逆順列範囲の例外テスト（"K:J" など）"""
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "Data"

        parser = SharePointExcelParser(self.mock_download_client)

        # 逆順序はValueErrorを発生させる
        with pytest.raises(ValueError) as exc_info:
            parser._normalize_column_range("K:J", ws)

        assert "無効なセル範囲" in str(exc_info.value)
        assert "K:J" in str(exc_info.value)

    def test_normalize_column_range_already_normalized(self):
        """すでに正規化済みの範囲はそのまま返すテスト"""
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        parser = SharePointExcelParser(self.mock_download_client)

        # すでに行番号付きの範囲はそのまま
        normalized = parser._normalize_column_range("A1:B10", ws)
        assert normalized == "A1:B10"

        # 単一セル
        normalized = parser._normalize_column_range("C5", ws)
        assert normalized == "C5"

        # 空文字列
        normalized = parser._normalize_column_range("", ws)
        assert normalized == ""

        # 空白のみ
        normalized = parser._normalize_column_range("  ", ws)
        assert normalized == "  "

    def test_no_duplicate_range_normalization(self):
        """
        セル範囲の正規化・拡張が重複して実行されないことを確認

        課題3-2の対応：_parse_sheetと_build_merged_cell_cacheで
        重複していた計算が1回のみになったことを検証
        """
        # テスト用Excelを作成（結合セルあり）
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "Header1"
        ws["B1"] = "Header2"
        ws["A2"] = "Data1"
        ws["B2"] = "Data2"
        ws["A3"] = "Data3"
        ws["B3"] = "Data4"

        # A1:B1を結合
        ws.merge_cells("A1:B1")

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        # モックの設定
        self.mock_download_client.download_file.return_value = excel_bytes.getvalue()

        parser = SharePointExcelParser(self.mock_download_client)

        # _normalize_column_rangeと_expand_axis_rangeの呼び出し回数をカウント
        with patch.object(
            parser, "_normalize_column_range", wraps=parser._normalize_column_range
        ) as mock_normalize, patch.object(
            parser, "_expand_axis_range", wraps=parser._expand_axis_range
        ) as mock_expand:
            # 列範囲指定で解析（expand_axis_range=Trueで拡張を有効化）
            result = parser.parse_to_json(
                "/test/file.xlsx", cell_range="A:B", expand_axis_range=True
            )
            result_data = json.loads(result)

            # 結果が正しいことを確認
            assert "sheets" in result_data
            assert len(result_data["sheets"]) == 1
            assert result_data["sheets"][0]["name"] == "TestSheet"

            # _normalize_column_rangeは1回だけ呼ばれる（重複なし）
            assert (
                mock_normalize.call_count == 1
            ), f"Expected 1 call, got {mock_normalize.call_count}"

            # _expand_axis_rangeは1回だけ呼ばれる（重複なし）
            assert (
                mock_expand.call_count == 1
            ), f"Expected 1 call, got {mock_expand.call_count}"

    def test_build_merged_cell_cache_with_effective_range(self):
        """
        _build_merged_cell_cacheにeffective_cell_rangeを渡した場合の動作確認

        計算済みの範囲を渡すことで、内部での重複計算が回避されることを検証
        """
        # テスト用Excelを作成（結合セルあり）
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "Merged"
        ws["A2"] = "Data1"
        ws["B2"] = "Data2"

        # A1:B1を結合
        ws.merge_cells("A1:B1")

        parser = SharePointExcelParser(self.mock_download_client)

        # effective_cell_rangeを渡して呼び出し
        merged_cell_map, merged_anchor_value_map, merged_ranges = (
            parser._build_merged_cell_cache(ws, effective_cell_range="A1:B2")
        )

        # 結合セル情報が正しく取得されることを確認
        assert merged_cell_map is not None
        assert merged_ranges is not None
        assert len(merged_ranges) == 1
        assert merged_ranges[0]["range"] == "A1:B1"

    def test_build_merged_cell_cache_without_effective_range(self):
        """
        _build_merged_cell_cacheにNoneを渡した場合の動作確認

        effective_cell_rangeがNoneの場合、sheet.dimensionsが使用されることを検証
        """
        # テスト用Excelを作成（結合セルあり）
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "Merged"
        ws["B1"] = "Header"
        ws["A2"] = "Data1"
        ws["B2"] = "Data2"

        # A1:B1を結合
        ws.merge_cells("A1:B1")

        parser = SharePointExcelParser(self.mock_download_client)

        # effective_cell_rangeにNoneを渡して呼び出し
        # sheet.dimensionsが使用される
        merged_cell_map, merged_anchor_value_map, merged_ranges = (
            parser._build_merged_cell_cache(ws, effective_cell_range=None)
        )

        # 結合セル情報が正しく取得されることを確認
        assert merged_cell_map is not None
        assert merged_ranges is not None
        assert len(merged_ranges) == 1
        assert merged_ranges[0]["range"] == "A1:B1"

    def test_range_normalization_integration(self):
        """
        セル範囲の正規化・拡張と結合セル処理の統合テスト

        列範囲指定（"A:B"）が正しく正規化・拡張され、
        結合セル情報も正しく取得されることを検証
        """
        # テスト用Excelを作成（結合セルあり）
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "Merged Header"
        ws["A2"] = "Data1"
        ws["B2"] = "Data2"
        ws["A3"] = "Data3"
        ws["B3"] = "Data4"

        # A1:B1を結合
        ws.merge_cells("A1:B1")

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        # モックの設定
        self.mock_download_client.download_file.return_value = excel_bytes.getvalue()

        parser = SharePointExcelParser(self.mock_download_client)

        # 列範囲指定で解析
        result = parser.parse_to_json("/test/file.xlsx", cell_range="A:B")
        result_data = json.loads(result)

        # 結果検証
        assert "sheets" in result_data
        assert len(result_data["sheets"]) == 1

        sheet_data = result_data["sheets"][0]
        assert sheet_data["name"] == "TestSheet"

        # requested_rangeとeffective_rangeが設定されている
        assert sheet_data["requested_range"] == "A:B"
        assert "effective_range" in sheet_data
        assert sheet_data["effective_range"].startswith("A1:B")

        # 結合セル情報が取得されている
        assert "merged_ranges" in sheet_data
        assert len(sheet_data["merged_ranges"]) == 1
        assert sheet_data["merged_ranges"][0]["range"] == "A1:B1"

        # データも正しく取得されている
        assert "rows" in sheet_data
        assert len(sheet_data["rows"]) > 0

    def _create_frozen_rows_excel(self, freeze_panes: str) -> bytes:
        """frozen_rows機能テスト用のExcelファイルを作成

        Args:
            freeze_panes: freeze_panesの指定（例: "A3"で2行固定）

        Returns:
            Excelファイルのバイトデータ
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # ヘッダー行（固定行）
        ws["A1"] = "名前"
        ws["B1"] = "年齢"
        ws["A2"] = "Name"
        ws["B2"] = "Age"

        # データ行
        ws["A4"] = "花子"
        ws["B4"] = 25
        ws["A5"] = "次郎"
        ws["B5"] = 30
        ws["A6"] = "三郎"
        ws["B6"] = 35

        # freeze_panesを設定
        ws.freeze_panes = freeze_panes

        # BytesIOに保存
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        return excel_bytes.getvalue()

    def test_include_frozen_rows_default_true(self):
        """デフォルト（include_frozen_rows=True）でヘッダーが自動追加されること"""
        # frozen_rows=2のExcelファイルを作成
        excel_bytes = self._create_frozen_rows_excel("A3")
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)

        # cell_range="A4:B6"を指定（データ範囲のみ）
        result_json = parser.parse_to_json(
            "test.xlsx", sheet_name="TestSheet", cell_range="A4:B6"
        )

        result = json.loads(result_json)
        sheet_data = result["sheets"][0]

        # frozen_rowsが2であることを確認
        assert sheet_data["frozen_rows"] == 2

        # rowsにヘッダー（A1:B2）+ データ（A4:B6）が含まれる
        assert len(sheet_data["rows"]) == 5  # 2 + 3

        # ヘッダー行（A1:B2）
        assert sheet_data["rows"][0][0]["value"] == "名前"
        assert sheet_data["rows"][0][0]["coordinate"] == "A1"
        assert sheet_data["rows"][1][0]["value"] == "Name"
        assert sheet_data["rows"][1][0]["coordinate"] == "A2"

        # データ行（A4:B6）
        assert sheet_data["rows"][2][0]["value"] == "花子"
        assert sheet_data["rows"][2][0]["coordinate"] == "A4"
        assert sheet_data["rows"][3][0]["value"] == "次郎"
        assert sheet_data["rows"][3][0]["coordinate"] == "A5"
        assert sheet_data["rows"][4][0]["value"] == "三郎"
        assert sheet_data["rows"][4][0]["coordinate"] == "A6"

    def test_include_frozen_rows_false(self):
        """include_frozen_rows=Falseでヘッダーが追加されないこと"""
        # frozen_rows=2のExcelファイルを作成
        excel_bytes = self._create_frozen_rows_excel("A3")
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)

        # include_frozen_rows=Falseを指定
        result_json = parser.parse_to_json(
            "test.xlsx",
            sheet_name="TestSheet",
            cell_range="A4:B6",
            include_frozen_rows=False,
        )

        result = json.loads(result_json)
        sheet_data = result["sheets"][0]

        # frozen_rowsが2であることを確認
        assert sheet_data["frozen_rows"] == 2

        # rowsにはデータ範囲（A4:B6）のみ
        assert len(sheet_data["rows"]) == 3

        # データ行のみ
        assert sheet_data["rows"][0][0]["value"] == "花子"
        assert sheet_data["rows"][0][0]["coordinate"] == "A4"
        assert sheet_data["rows"][1][0]["value"] == "次郎"
        assert sheet_data["rows"][1][0]["coordinate"] == "A5"
        assert sheet_data["rows"][2][0]["value"] == "三郎"
        assert sheet_data["rows"][2][0]["coordinate"] == "A6"

    def test_include_frozen_rows_no_frozen_rows(self):
        """frozen_rows=0の場合はヘッダー追加なし"""
        # frozen_panesなしのExcelファイルを作成
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A4"] = "花子"
        ws["B4"] = 25
        ws["A5"] = "次郎"
        ws["B5"] = 30

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        self.mock_download_client.download_file.return_value = excel_bytes.getvalue()

        parser = SharePointExcelParser(self.mock_download_client)

        result_json = parser.parse_to_json(
            "test.xlsx", sheet_name="TestSheet", cell_range="A4:B5"
        )

        result = json.loads(result_json)
        sheet_data = result["sheets"][0]

        # frozen_rowsが0
        assert sheet_data["frozen_rows"] == 0

        # header_detectionが付与される
        assert "header_detection" in sheet_data
        assert sheet_data["header_detection"]["status"] == "no_frozen_rows"
        assert sheet_data["header_detection"]["frozen_rows"] == 0
        assert len(sheet_data["header_detection"]["suggestions"]) == 2

        # rowsにはデータ範囲のみ
        assert len(sheet_data["rows"]) == 2

    def test_include_frozen_rows_range_includes_headers(self):
        """cell_rangeが既にヘッダーを含む場合は重複回避"""
        # frozen_rows=2のExcelファイルを作成
        excel_bytes = self._create_frozen_rows_excel("A3")
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)

        # cell_range="A1:B6"（ヘッダーを含む）
        result_json = parser.parse_to_json(
            "test.xlsx", sheet_name="TestSheet", cell_range="A1:B6"
        )

        result = json.loads(result_json)
        sheet_data = result["sheets"][0]

        # frozen_rowsが2
        assert sheet_data["frozen_rows"] == 2

        # rowsには全データ（重複なし）
        # A1, A2は1回ずつ、A4, A5, A6も1回ずつ
        rows = sheet_data["rows"]

        # A1が最初に1回だけ登場
        a1_count = sum(
            1 for row in rows for cell in row if cell["coordinate"] == "A1"
        )
        assert a1_count == 1

        # A2も1回だけ登場
        a2_count = sum(
            1 for row in rows for cell in row if cell["coordinate"] == "A2"
        )
        assert a2_count == 1

    def test_include_frozen_rows_single_cell(self):
        """単一セル指定時にデフォルト（expand_axis_range=False）では軸拡張せずヘッダー + 指定セルが追加される"""
        # frozen_rows=2のExcelファイルを作成
        excel_bytes = self._create_frozen_rows_excel("A3")
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)

        # cell_range="B5"（単一セル）
        # expand_axis_range=False（デフォルト）なので拡張されない
        # include_frozen_rows=TrueでB1:B2（ヘッダー）が追加される
        result_json = parser.parse_to_json(
            "test.xlsx", sheet_name="TestSheet", cell_range="B5"
        )

        result = json.loads(result_json)
        sheet_data = result["sheets"][0]

        # frozen_rowsが2
        assert sheet_data["frozen_rows"] == 2

        # B5のみ + ヘッダーB1:B2 → 3行
        assert len(sheet_data["rows"]) == 3

        # B1, B2（ヘッダー）, B5（指定セル）の順
        assert sheet_data["rows"][0][0]["coordinate"] == "B1"
        assert sheet_data["rows"][1][0]["coordinate"] == "B2"
        assert sheet_data["rows"][2][0]["coordinate"] == "B5"

    def test_expand_axis_range_true_single_column(self):
        """expand_axis_range=Trueで単一列が行1まで拡張されること"""
        excel_bytes = self._create_frozen_rows_excel("A3")
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)

        # cell_range="B4:B6"（単一列の部分範囲）
        # expand_axis_range=Trueなので B1:B6 に拡張される
        result_json = parser.parse_to_json(
            "test.xlsx",
            sheet_name="TestSheet",
            cell_range="B4:B6",
            expand_axis_range=True,
        )

        result = json.loads(result_json)
        sheet_data = result["sheets"][0]

        # B1:B6に拡張（6行）、ヘッダーB1:B2は範囲に含まれるため重複追加なし
        assert len(sheet_data["rows"]) == 6
        assert sheet_data["rows"][0][0]["coordinate"] == "B1"
        assert sheet_data["rows"][5][0]["coordinate"] == "B6"

    def test_expand_axis_range_false_default(self):
        """expand_axis_range=False（デフォルト）で拡張されないこと"""
        excel_bytes = self._create_frozen_rows_excel("A3")
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)

        # cell_range="B4:B6"（単一列の部分範囲）
        # expand_axis_range=False（デフォルト）なので拡張されない
        result_json = parser.parse_to_json(
            "test.xlsx",
            sheet_name="TestSheet",
            cell_range="B4:B6",
        )

        result = json.loads(result_json)
        sheet_data = result["sheets"][0]

        # B4:B6のみ + ヘッダーB1:B2（include_frozen_rows=Trueがデフォルト）
        assert len(sheet_data["rows"]) == 5
        assert sheet_data["rows"][0][0]["coordinate"] == "B1"
        assert sheet_data["rows"][1][0]["coordinate"] == "B2"
        assert sheet_data["rows"][2][0]["coordinate"] == "B4"
        assert sheet_data["rows"][3][0]["coordinate"] == "B5"
        assert sheet_data["rows"][4][0]["coordinate"] == "B6"

    def test_expand_axis_range_with_include_frozen_rows(self):
        """expand_axis_range=TrueとFalseでinclude_frozen_rowsとの組み合わせ"""
        excel_bytes = self._create_frozen_rows_excel("A3")
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)

        # expand_axis_range=True, include_frozen_rows=False
        # B4:B6 → B1:B6に拡張されるが、ヘッダー自動追加はなし
        result_json = parser.parse_to_json(
            "test.xlsx",
            sheet_name="TestSheet",
            cell_range="B4:B6",
            include_frozen_rows=False,
            expand_axis_range=True,
        )

        result = json.loads(result_json)
        sheet_data = result["sheets"][0]

        # B1:B6に拡張（6行）、include_frozen_rows=Falseなのでヘッダー追加なし
        assert len(sheet_data["rows"]) == 6
        assert sheet_data["rows"][0][0]["coordinate"] == "B1"
        assert sheet_data["rows"][5][0]["coordinate"] == "B6"

    def test_expand_axis_range_true_single_cell(self):
        """expand_axis_range=Trueで単一セルが拡張されること"""
        excel_bytes = self._create_frozen_rows_excel("A3")
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)

        # cell_range="B5"（単一セル）
        # expand_axis_range=Trueなので B1:B5 に拡張される
        result_json = parser.parse_to_json(
            "test.xlsx",
            sheet_name="TestSheet",
            cell_range="B5",
            expand_axis_range=True,
        )

        result = json.loads(result_json)
        sheet_data = result["sheets"][0]

        # B1:B5に拡張（5行）、ヘッダーB1:B2は範囲に含まれるため重複追加なし
        assert len(sheet_data["rows"]) == 5
        assert sheet_data["rows"][0][0]["coordinate"] == "B1"
        assert sheet_data["rows"][4][0]["coordinate"] == "B5"

    def test_include_frozen_rows_without_cell_range(self):
        """cell_range未指定時はヘッダー追加不要（全シート取得）"""
        # frozen_rows=2のExcelファイルを作成
        excel_bytes = self._create_frozen_rows_excel("A3")
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)

        # cell_rangeなし（全シート取得）
        result_json = parser.parse_to_json("test.xlsx", sheet_name="TestSheet")

        result = json.loads(result_json)
        sheet_data = result["sheets"][0]

        # frozen_rowsが2
        assert sheet_data["frozen_rows"] == 2

        # rowsには全データ（A1～A6）
        # requested_rangeは設定されていない
        assert "requested_range" not in sheet_data

        # cell_range未指定なので header_detection は付与されない
        assert "header_detection" not in sheet_data

    def test_header_detection_warning_frozen_rows_zero_with_cell_range(self):
        """frozen_rows=0 かつ cell_range指定時に header_detection が付与される"""
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "Header1"
        ws["B1"] = "Header2"
        ws["A5"] = "Data1"
        ws["B5"] = "Data2"

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        self.mock_download_client.download_file.return_value = excel_bytes.getvalue()

        parser = SharePointExcelParser(self.mock_download_client)

        result_json = parser.parse_to_json(
            "test.xlsx", sheet_name="TestSheet", cell_range="A5:B5"
        )

        result = json.loads(result_json)
        sheet_data = result["sheets"][0]

        # frozen_rows=0
        assert sheet_data["frozen_rows"] == 0

        # header_detection が付与される
        assert "header_detection" in sheet_data
        assert sheet_data["header_detection"]["status"] == "no_frozen_rows"
        assert sheet_data["header_detection"]["frozen_rows"] == 0
        assert "note" in sheet_data["header_detection"]
        assert len(sheet_data["header_detection"]["suggestions"]) == 2

    def test_header_detection_not_added_when_frozen_rows_exist(self):
        """frozen_rows > 0 の場合は header_detection が付与されない"""
        excel_bytes = self._create_frozen_rows_excel("A3")
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)

        result_json = parser.parse_to_json(
            "test.xlsx", sheet_name="TestSheet", cell_range="A4:B5"
        )

        result = json.loads(result_json)
        sheet_data = result["sheets"][0]

        # frozen_rows=2
        assert sheet_data["frozen_rows"] == 2

        # header_detection は付与されない（frozen_rows > 0）
        assert "header_detection" not in sheet_data

    def test_header_detection_not_added_when_expand_axis_range_true(self):
        """expand_axis_range=True の場合は header_detection が付与されない"""
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "Header1"
        ws["B1"] = "Header2"
        ws["A5"] = "Data1"
        ws["B5"] = "Data2"

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        self.mock_download_client.download_file.return_value = excel_bytes.getvalue()

        parser = SharePointExcelParser(self.mock_download_client)

        # frozen_rows=0 だが expand_axis_range=True なので警告不要
        result_json = parser.parse_to_json(
            "test.xlsx", sheet_name="TestSheet", cell_range="A5:B5", expand_axis_range=True
        )

        result = json.loads(result_json)
        sheet_data = result["sheets"][0]

        # frozen_rows=0
        assert sheet_data["frozen_rows"] == 0

        # header_detection は付与されない（expand_axis_range=True で1行目が含まれる）
        assert "header_detection" not in sheet_data

    def test_include_frozen_rows_with_merged_cells(self):
        """マージセルとの統合が正しく機能すること"""
        # frozen_rows=2、マージセル付きのExcelファイルを作成
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # ヘッダー行（マージセル）
        ws["A1"] = "名前"
        ws.merge_cells("A1:B1")  # ヘッダーをマージ

        ws["A2"] = "Name"
        ws["B2"] = "Age"

        # データ行
        ws["A4"] = "花子"
        ws["B4"] = 25
        ws["A5"] = "次郎"
        ws["B5"] = 30

        # freeze_panesを設定
        ws.freeze_panes = "A3"

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        self.mock_download_client.download_file.return_value = excel_bytes.getvalue()

        parser = SharePointExcelParser(self.mock_download_client)

        # cell_range="A4:B5"を指定
        result_json = parser.parse_to_json(
            "test.xlsx", sheet_name="TestSheet", cell_range="A4:B5"
        )

        result = json.loads(result_json)
        sheet_data = result["sheets"][0]

        # frozen_rowsが2
        assert sheet_data["frozen_rows"] == 2

        # マージセル情報が含まれる
        assert "merged_ranges" in sheet_data
        assert len(sheet_data["merged_ranges"]) == 1
        assert sheet_data["merged_ranges"][0]["range"] == "A1:B1"

        # rowsにはヘッダー（A1:B2）+ データ（A4:B5）
        assert len(sheet_data["rows"]) == 4

        # ヘッダー行（A1はマージセル）
        assert sheet_data["rows"][0][0]["value"] == "名前"
        assert sheet_data["rows"][0][0]["coordinate"] == "A1"
        assert sheet_data["rows"][0][0]["merged"]["is_top_left"] is True
        assert sheet_data["rows"][0][0]["merged"]["range"] == "A1:B1"

    def test_frozen_rows_dos_mitigation_exceeds_limit(self):
        """frozen_rowsが上限を超えた場合は0にリセットされて処理が続行されること（DoS対策）"""
        # 異常に大きなfrozen_rowsを持つExcelファイルを作成
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "Header"
        ws["A2"] = "Data"

        # 異常に大きなfrozen_rowsを設定（openpyxl内部でpane.ySplitに直接設定）
        pane = Pane()
        pane.ySplit = 200  # デフォルト上限(100)を超える値
        pane.xSplit = 0
        pane.topLeftCell = "A201"
        pane.state = "frozen"

        sheet_view = SheetView(pane=pane)
        ws.views.sheetView[0] = sheet_view

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        self.mock_download_client.download_file.return_value = excel_bytes.getvalue()

        parser = SharePointExcelParser(self.mock_download_client)

        # frozen_rowsが上限を超えているが、リセットされて処理が続行される
        result_json = parser.parse_to_json("test.xlsx")
        result = json.loads(result_json)

        # frozen_rowsが0にリセットされている
        sheet_data = result["sheets"][0]
        assert sheet_data["frozen_rows"] == 0
        assert sheet_data["frozen_cols"] == 0
        # freeze_panes情報は含まれない
        assert "freeze_panes" not in sheet_data

    @patch("src.sharepoint_excel.config.excel_max_frozen_rows", 50)
    def test_frozen_rows_dos_mitigation_within_limit(self):
        """frozen_rowsが上限以内の場合は正常に処理されること"""
        # 上限以内のfrozen_rowsを持つExcelファイルを作成
        excel_bytes = self._create_frozen_rows_excel("A3")  # frozen_rows=2
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)

        # 正常に処理される（エラーが発生しない）
        result_json = parser.parse_to_json("test.xlsx")
        result = json.loads(result_json)

        # frozen_rowsが正しく取得されている
        sheet_data = result["sheets"][0]
        assert sheet_data["frozen_rows"] == 2

    @patch("src.sharepoint_excel.config.excel_max_frozen_rows", 150)
    def test_frozen_rows_dos_mitigation_custom_limit(self):
        """カスタム上限値が正しく適用されること"""
        # カスタム上限(150)以内のfrozen_rowsを持つExcelファイルを作成
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "Header"

        # frozen_rows=120を設定（カスタム上限150以内）
        pane = Pane()
        pane.ySplit = 120
        pane.xSplit = 0
        pane.topLeftCell = "A121"
        pane.state = "frozen"

        sheet_view = SheetView(pane=pane)
        ws.views.sheetView[0] = sheet_view

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        self.mock_download_client.download_file.return_value = excel_bytes.getvalue()

        parser = SharePointExcelParser(self.mock_download_client)

        # カスタム上限(150)以内なので正常に処理される
        result_json = parser.parse_to_json("test.xlsx")
        result = json.loads(result_json)

        sheet_data = result["sheets"][0]
        assert sheet_data["frozen_rows"] == 120

    def test_parse_with_cell_styles_disabled(self):
        """include_cell_styles=False（デフォルト）でスタイル情報が含まれないこと"""
        excel_bytes = self._create_formatted_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        # デフォルト（include_cell_styles=False）で解析
        result_json = parser.parse_to_json("/test/formatted.xlsx")

        result = json.loads(result_json)
        cell = result["sheets"][0]["rows"][0][0]

        # value と coordinate は含まれる
        assert "value" in cell
        assert "coordinate" in cell
        assert cell["value"] == "Name"

        # スタイル情報は含まれない
        assert "fill" not in cell
        assert "width" not in cell
        assert "height" not in cell

    def test_parse_with_cell_styles_enabled(self):
        """include_cell_styles=Trueでスタイル情報が含まれること"""
        excel_bytes = self._create_formatted_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        # include_cell_styles=Trueで解析
        result_json = parser.parse_to_json(
            "/test/formatted.xlsx", include_cell_styles=True
        )

        result = json.loads(result_json)
        cell = result["sheets"][0]["rows"][0][0]

        # 基本情報は含まれる
        assert cell["value"] == "Name"
        assert cell["coordinate"] == "A1"

        # 背景色情報が含まれる
        assert "fill" in cell
        assert cell["fill"]["pattern_type"] == "solid"
        assert "fg_color" in cell["fill"]
        # 16進数形式であることを確認
        if cell["fill"]["fg_color"]:
            assert cell["fill"]["fg_color"].startswith("#")

    def test_cell_styles_with_width_height(self):
        """列幅・行高さが設定されたセルでwidth/heightが正しく返されること"""
        wb = Workbook()
        ws = wb.active
        ws.title = "SizedSheet"

        # セルにデータを設定
        ws["A1"] = "Wide Column"
        ws["A2"] = "Tall Row"

        # 列幅と行高さを設定
        ws.column_dimensions["A"].width = 30
        ws.row_dimensions[2].height = 50

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        self.mock_download_client.download_file.return_value = excel_bytes.getvalue()

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.parse_to_json("/test/sized.xlsx", include_cell_styles=True)

        result = json.loads(result_json)
        rows = result["sheets"][0]["rows"]

        # A1セル: 列幅が設定されている
        cell_a1 = rows[0][0]
        assert cell_a1["coordinate"] == "A1"
        assert "width" in cell_a1
        assert cell_a1["width"] == 30

        # A2セル: 列幅と行高さが設定されている
        cell_a2 = rows[1][0]
        assert cell_a2["coordinate"] == "A2"
        assert "width" in cell_a2
        assert cell_a2["width"] == 30
        assert "height" in cell_a2
        assert cell_a2["height"] == 50

    def test_cell_styles_none_when_not_set(self):
        """スタイルが設定されていないセルではfillなどが含まれないこと"""
        # シンプルなExcelファイルを作成（スタイルなし）
        excel_bytes = self._create_test_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        # include_cell_styles=Trueで解析
        result_json = parser.parse_to_json(
            "/test/file.xlsx", include_cell_styles=True
        )

        result = json.loads(result_json)
        cell = result["sheets"][0]["rows"][0][0]

        # 基本情報は含まれる
        assert cell["value"] == "Name"
        assert cell["coordinate"] == "A1"

        # スタイルが設定されていないので、fillやwidth/heightは含まれない
        assert "fill" not in cell
        assert "width" not in cell
        assert "height" not in cell

    def test_omit_fixed_value_fields(self):
        """固定値フィールドが省略されること"""
        excel_bytes = self._create_test_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        result_json = parser.parse_to_json("/test/file.xlsx")

        result = json.loads(result_json)

        # トップレベルの固定値フィールドは省略
        assert "response_kind" not in result
        assert "data_included" not in result

        # シートレベルの固定値フィールドは省略
        sheet = result["sheets"][0]
        assert "purpose" not in sheet
        assert "data_included" not in sheet

    def test_omit_null_fields(self):
        """nullフィールドが省略されること"""
        excel_bytes = self._create_test_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        # sheet_name/cell_rangeを指定しない
        result_json = parser.parse_to_json("/test/file.xlsx")

        result = json.loads(result_json)

        # nullフィールドは省略
        assert "requested_sheet" not in result
        assert "requested_range" not in result

    def test_include_non_null_fields(self):
        """null以外のフィールドが含まれること"""
        excel_bytes = self._create_multi_sheet_excel()
        self.mock_download_client.download_file.return_value = excel_bytes

        parser = SharePointExcelParser(self.mock_download_client)
        # sheet_name/cell_rangeを指定
        result_json = parser.parse_to_json(
            "/test/file.xlsx", sheet_name="Sheet1", cell_range="A1:B2"
        )

        result = json.loads(result_json)

        # 値があるフィールドは含まれる
        assert result["requested_sheet"] == "Sheet1"
        assert result["requested_range"] == "A1:B2"

    def test_omit_null_dimensions(self):
        """dimensions=nullが省略されること"""
        parser = SharePointExcelParser(self.mock_download_client)

        # モックを使用してsheet.dimensionsをNoneに設定
        with patch("src.sharepoint_excel.load_workbook") as mock_load:
            mock_wb = Mock()
            mock_sheet = Mock()
            mock_sheet.title = "EmptySheet"
            mock_sheet.dimensions = None  # dimensionsをNoneに設定
            mock_sheet.freeze_panes = None  # freeze_panesもNone
            # sheet_viewを設定して_get_frozen_panes()での警告ログを防ぐ
            mock_sheet.sheet_view = Mock(pane=None)
            # merged_cellsはranges属性を持つオブジェクト
            mock_merged_cells = Mock()
            mock_merged_cells.ranges = []
            mock_sheet.merged_cells = mock_merged_cells
            mock_sheet.iter_rows = Mock(return_value=[])  # 行なし
            mock_sheet.max_row = 0
            mock_sheet.max_column = 0
            mock_wb.sheetnames = ["EmptySheet"]
            mock_wb.__getitem__ = Mock(return_value=mock_sheet)
            mock_wb.close = Mock()
            mock_load.return_value = mock_wb

            # load_workbookがモックされているため、download_fileの戻り値は実際には使われない
            self.mock_download_client.download_file.return_value = b""
            result_json = parser.parse_to_json("/test/empty.xlsx")

        result = json.loads(result_json)
        sheet = result["sheets"][0]

        # dimensionsがNoneの場合は省略
        assert "dimensions" not in sheet
