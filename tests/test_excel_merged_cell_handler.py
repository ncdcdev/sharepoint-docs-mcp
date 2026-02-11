"""
ExcelMergedCellHandlerのテスト
"""

from io import BytesIO

from openpyxl import Workbook

from src.excel import ExcelMergedCellHandler


class TestExcelMergedCellHandler:
    """ExcelMergedCellHandler（マージセル処理）のテスト"""

    def _create_test_workbook_with_merged_cells(self) -> Workbook:
        """結合セルを含むテスト用Workbookを作成"""
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # 通常のセル
        ws["A1"] = "Header1"
        ws["B1"] = "Header2"
        ws["A2"] = "Data1"
        ws["B2"] = "Data2"

        # 結合セル（A3:B3）
        ws.merge_cells("A3:B3")
        ws["A3"] = "Merged Cell"

        # 結合セル（A4:B5）
        ws.merge_cells("A4:B5")
        ws["A4"] = "Large Merged"

        return wb

    def _simple_serializer(self, value):
        """テスト用の簡易シリアライザー"""
        if value is None:
            return None
        if isinstance(value, (str, int, float, bool)):
            return value
        return str(value)

    def test_build_merged_cell_cache_basic(self):
        """基本的なマージセル情報の構築テスト"""
        wb = self._create_test_workbook_with_merged_cells()
        ws = wb.active

        # 全範囲を対象にキャッシュ構築
        merged_cell_map, merged_anchor_value_map, merged_ranges = (
            ExcelMergedCellHandler.build_merged_cell_cache(
                ws, "A1:B5", self._simple_serializer
            )
        )

        # 2つの結合セルが検出されること
        assert len(merged_ranges) == 2

        # 結合範囲が正しいこと
        ranges = [mr["range"] for mr in merged_ranges]
        assert "A3:B3" in ranges
        assert "A4:B5" in ranges

        # アンカー値が正しいこと
        assert merged_anchor_value_map["A3:B3"] == "Merged Cell"
        assert merged_anchor_value_map["A4:B5"] == "Large Merged"

        # マージセルマップが正しいこと
        assert merged_cell_map["A3"] == "A3:B3"
        assert merged_cell_map["B3"] == "A3:B3"
        assert merged_cell_map["A4"] == "A4:B5"
        assert merged_cell_map["B4"] == "A4:B5"
        assert merged_cell_map["A5"] == "A4:B5"
        assert merged_cell_map["B5"] == "A4:B5"

    def test_build_merged_cell_cache_partial_range(self):
        """部分範囲でのマージセル情報の構築テスト（部分展開）"""
        wb = self._create_test_workbook_with_merged_cells()
        ws = wb.active

        # A3:B3のみを対象
        merged_cell_map, merged_anchor_value_map, merged_ranges = (
            ExcelMergedCellHandler.build_merged_cell_cache(
                ws, "A3:B3", self._simple_serializer
            )
        )

        # 1つの結合セルのみ検出されること
        assert len(merged_ranges) == 1
        assert merged_ranges[0]["range"] == "A3:B3"

        # A4:B5は範囲外なのでマップに含まれない
        assert "A4" not in merged_cell_map

    def test_build_merged_cell_cache_no_merged_cells(self):
        """結合セルがない場合のテスト"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Data"

        merged_cell_map, merged_anchor_value_map, merged_ranges = (
            ExcelMergedCellHandler.build_merged_cell_cache(
                ws, "A1:B2", self._simple_serializer
            )
        )

        # 空の結果が返ること
        assert merged_cell_map is None
        assert merged_anchor_value_map is None
        assert merged_ranges == []

    def test_build_merged_cell_cache_none_range(self):
        """effective_cell_range=Noneの場合はsheet.dimensionsが使用されること"""
        wb = self._create_test_workbook_with_merged_cells()
        ws = wb.active

        # effective_cell_rangeをNoneに設定
        merged_cell_map, merged_anchor_value_map, merged_ranges = (
            ExcelMergedCellHandler.build_merged_cell_cache(
                ws, None, self._simple_serializer
            )
        )

        # sheet.dimensions全体が対象になるので、2つの結合セルが検出される
        assert len(merged_ranges) == 2

    def test_build_merged_cell_cache_anchor_value_in_non_topleft(self):
        """openpyxlで新規作成時はマージセル内の値が失われることを確認"""
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # マージセル作成前に値を設定
        ws["B2"] = "Anchor Value"
        # 結合セル（A1:B2）を作成
        ws.merge_cells("A1:B2")

        # openpyxlで新規作成した場合、マージ後は左上以外の値が失われる
        # （実際のExcelファイルを読み込む場合は異なる動作になる）
        merged_cell_map, merged_anchor_value_map, merged_ranges = (
            ExcelMergedCellHandler.build_merged_cell_cache(
                ws, "A1:B2", self._simple_serializer
            )
        )

        assert len(merged_ranges) == 1
        assert merged_ranges[0]["range"] == "A1:B2"
        # openpyxlで新規作成時は左上が空、値も失われる
        assert merged_ranges[0]["anchor"]["coordinate"] == "A1"
        assert merged_ranges[0]["anchor"]["value"] is None
        assert merged_anchor_value_map["A1:B2"] is None

    def test_build_merged_cell_cache_all_empty_cells(self):
        """結合セル内の全てのセルが空の場合のテスト"""
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # 結合セル（A1:B2）で全て空
        ws.merge_cells("A1:B2")
        # 値を設定しない

        merged_cell_map, merged_anchor_value_map, merged_ranges = (
            ExcelMergedCellHandler.build_merged_cell_cache(
                ws, "A1:B2", self._simple_serializer
            )
        )

        assert len(merged_ranges) == 1
        assert merged_ranges[0]["range"] == "A1:B2"
        # 全て空の場合は左上がアンカー、値はNone
        assert merged_ranges[0]["anchor"]["coordinate"] == "A1"
        assert merged_ranges[0]["anchor"]["value"] is None
        assert merged_anchor_value_map["A1:B2"] is None

    def test_build_merged_cell_cache_intersection_only(self):
        """結合範囲と対象範囲の交差部分のみがマップに含まれること"""
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # 結合セル（A1:D4）を作成
        ws.merge_cells("A1:D4")
        ws["A1"] = "Merged"

        # 対象範囲をB2:C3に限定（結合範囲の一部）
        merged_cell_map, merged_anchor_value_map, merged_ranges = (
            ExcelMergedCellHandler.build_merged_cell_cache(
                ws, "B2:C3", self._simple_serializer
            )
        )

        # 結合情報は取得されるが、マップには交差部分のみ
        assert len(merged_ranges) == 1
        assert merged_ranges[0]["range"] == "A1:D4"

        # 交差部分（B2, B3, C2, C3）のみマップに含まれる
        assert "B2" in merged_cell_map
        assert "B3" in merged_cell_map
        assert "C2" in merged_cell_map
        assert "C3" in merged_cell_map

        # 交差していない部分はマップに含まれない
        assert "A1" not in merged_cell_map
        assert "D4" not in merged_cell_map

    def test_build_merged_cell_cache_no_intersection(self):
        """結合範囲と対象範囲が交差しない場合は検出されないこと"""
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # 結合セル（A1:B2）を作成
        ws.merge_cells("A1:B2")
        ws["A1"] = "Merged"

        # 対象範囲をD4:E5に限定（結合範囲と交差しない）
        merged_cell_map, merged_anchor_value_map, merged_ranges = (
            ExcelMergedCellHandler.build_merged_cell_cache(
                ws, "D4:E5", self._simple_serializer
            )
        )

        # 交差しないので結合セルは検出されない
        assert merged_cell_map is None
        assert merged_anchor_value_map is None
        assert merged_ranges == []

    def test_build_merged_cell_cache_with_dollar_signs(self):
        """$記号付きの範囲が正しく処理されること"""
        wb = self._create_test_workbook_with_merged_cells()
        ws = wb.active

        # $記号付きの範囲
        merged_cell_map, merged_anchor_value_map, merged_ranges = (
            ExcelMergedCellHandler.build_merged_cell_cache(
                ws, "$A$1:$B$5", self._simple_serializer
            )
        )

        # 正しく処理されること
        assert len(merged_ranges) == 2

    def test_build_merged_cell_cache_single_cell_range(self):
        """単一セル範囲の場合のテスト"""
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # 結合セル（A1:B2）を作成
        ws.merge_cells("A1:B2")
        ws["A1"] = "Merged"

        # 単一セル範囲（A1）
        merged_cell_map, merged_anchor_value_map, merged_ranges = (
            ExcelMergedCellHandler.build_merged_cell_cache(
                ws, "A1", self._simple_serializer
            )
        )

        # A1を含む結合セルが検出される
        assert len(merged_ranges) == 1
        assert merged_ranges[0]["range"] == "A1:B2"

        # マップにはA1のみ（交差部分）
        assert "A1" in merged_cell_map
        assert "B2" not in merged_cell_map
