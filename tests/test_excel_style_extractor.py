"""
ExcelStyleExtractorのテスト
"""

from io import BytesIO
from unittest.mock import Mock

from openpyxl import Workbook
from openpyxl.styles import Color, Font, PatternFill

from src.excel import ExcelStyleExtractor


class TestExcelStyleExtractor:
    """ExcelStyleExtractor（スタイル抽出）のテスト"""

    # color_to_hex のテスト

    def test_color_to_hex_rgb(self):
        """RGB色が16進数に変換されること"""
        # RGB形式のColorオブジェクト
        color = Color(rgb="FFFF0000")  # 赤
        result = ExcelStyleExtractor.color_to_hex(color)
        assert result == "#FF0000"

    def test_color_to_hex_rgb_with_alpha(self):
        """アルファ値を含むRGB色が正しく変換されること"""
        # ARGB形式（最初の2桁はアルファ値）
        color = Color(rgb="80FF0000")  # 透明度50%の赤
        result = ExcelStyleExtractor.color_to_hex(color)
        # 下6桁を取得
        assert result == "#FF0000"

    def test_color_to_hex_theme(self):
        """テーマ色が文字列表現に変換されること"""
        color = Color(theme=1)
        result = ExcelStyleExtractor.color_to_hex(color)
        assert result == "theme_1"

    def test_color_to_hex_none(self):
        """Noneの場合はNoneが返ること"""
        result = ExcelStyleExtractor.color_to_hex(None)
        assert result is None

    # build_cell_size_cache のテスト

    def test_build_cell_size_cache_with_sizes(self):
        """列幅・行高さが設定されている場合"""
        wb = Workbook()
        ws = wb.active

        # 列幅を設定
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30

        # 行高さを設定
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 35

        col_widths, row_heights = ExcelStyleExtractor.build_cell_size_cache(ws)

        assert col_widths == {"A": 20, "B": 30}
        assert row_heights == {1: 25, 2: 35}

    def test_build_cell_size_cache_empty(self):
        """列幅・行高さが設定されていない場合"""
        wb = Workbook()
        ws = wb.active

        col_widths, row_heights = ExcelStyleExtractor.build_cell_size_cache(ws)

        assert col_widths == {}
        assert row_heights == {}

    def test_build_cell_size_cache_partial(self):
        """一部の列・行のみサイズが設定されている場合"""
        wb = Workbook()
        ws = wb.active

        # 一部の列幅のみ設定
        ws.column_dimensions["A"].width = 20
        # Bは設定しない

        # 一部の行高さのみ設定
        ws.row_dimensions[1].height = 25
        # 2は設定しない

        col_widths, row_heights = ExcelStyleExtractor.build_cell_size_cache(ws)

        assert col_widths == {"A": 20}
        assert row_heights == {1: 25}

    # extract_cell_styles のテスト

    def test_extract_cell_styles_with_fill(self):
        """背景色が設定されているセル"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Test"
        ws["A1"].fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        styles = ExcelStyleExtractor.extract_cell_styles(ws["A1"], None, None)

        assert "fill" in styles
        assert styles["fill"]["pattern_type"] == "solid"
        assert "fg_color" in styles["fill"]
        assert styles["fill"]["fg_color"] == "#FFFF00"

    def test_extract_cell_styles_with_size(self):
        """列幅・行高さが設定されているセル"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Test"
        ws.column_dimensions["A"].width = 30
        ws.row_dimensions[1].height = 50

        # キャッシュを作成
        col_widths, row_heights = ExcelStyleExtractor.build_cell_size_cache(ws)

        styles = ExcelStyleExtractor.extract_cell_styles(
            ws["A1"], col_widths, row_heights
        )

        assert "width" in styles
        assert styles["width"] == 30
        assert "height" in styles
        assert styles["height"] == 50

    def test_extract_cell_styles_no_styles(self):
        """スタイルが設定されていないセル"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Test"

        styles = ExcelStyleExtractor.extract_cell_styles(ws["A1"], None, None)

        # スタイル情報が含まれない
        assert "fill" not in styles
        assert "width" not in styles
        assert "height" not in styles
        assert styles == {}

    def test_extract_cell_styles_with_all_styles(self):
        """全てのスタイルが設定されているセル"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Test"
        ws["A1"].fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        ws.column_dimensions["A"].width = 30
        ws.row_dimensions[1].height = 50

        # キャッシュを作成
        col_widths, row_heights = ExcelStyleExtractor.build_cell_size_cache(ws)

        styles = ExcelStyleExtractor.extract_cell_styles(
            ws["A1"], col_widths, row_heights
        )

        assert "fill" in styles
        assert "width" in styles
        assert "height" in styles

    def test_extract_cell_styles_merged_cell(self):
        """MergedCellの場合も安全に処理されること"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Test"
        ws.merge_cells("A1:B2")

        # B2はMergedCellになる
        merged_cell = ws["B2"]

        # MergedCellにはcolumn_letter/row属性がないため、安全に処理される
        styles = ExcelStyleExtractor.extract_cell_styles(merged_cell, None, None)

        # エラーが発生せず、空のdictが返る
        assert isinstance(styles, dict)

    def test_extract_cell_styles_with_cache_miss(self):
        """キャッシュにない列・行の場合"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Test"
        ws.column_dimensions["B"].width = 30  # Bの幅を設定（Aではない）
        ws.row_dimensions[2].height = 50  # 2行目の高さを設定（1行目ではない）

        # キャッシュを作成
        col_widths, row_heights = ExcelStyleExtractor.build_cell_size_cache(ws)

        styles = ExcelStyleExtractor.extract_cell_styles(
            ws["A1"], col_widths, row_heights
        )

        # キャッシュにないのでwidth/heightは含まれない
        assert "width" not in styles
        assert "height" not in styles

    def test_extract_cell_styles_with_none_cache(self):
        """キャッシュがNoneの場合"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Test"
        ws.column_dimensions["A"].width = 30
        ws.row_dimensions[1].height = 50

        # キャッシュをNoneで渡す
        styles = ExcelStyleExtractor.extract_cell_styles(ws["A1"], None, None)

        # キャッシュがないのでwidth/heightは含まれない
        assert "width" not in styles
        assert "height" not in styles
